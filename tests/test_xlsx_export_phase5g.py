import io
import re

from openpyxl import load_workbook

import app


def _ids_from_html(body):
    return {int(value) for value in re.findall(rb'data-ticket-id="(\d+)"', body)}


def _ids_from_xlsx(response):
    sheet = load_workbook(io.BytesIO(response.data), read_only=True)["Current View"]
    return [row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None]


def test_export_current_view_matches_html_and_has_required_workbook_shape(client):
    query = "mode=closed&photo_video_only=1&missing_tags=1&workflow_tab=main"
    html = client.get("/queue?" + query)
    exported = client.get("/queue/export.xlsx?" + query)

    assert html.status_code == 200
    assert exported.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in exported.content_type
    assert "attachment" in exported.headers["Content-Disposition"]
    assert exported.headers["Content-Disposition"].split("=")[-1].endswith(".xlsx")
    assert _ids_from_xlsx(exported) == sorted(_ids_from_html(html.data))

    workbook = load_workbook(io.BytesIO(exported.data))
    sheet = workbook["Current View"]
    assert [cell.value for cell in sheet[1]] == [
        "Ticket #", "Subject", "Status", "Priority", "Updated", "Created",
        "Tags", "Review Result", "Freshdesk URL",
    ]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == f"A1:I{sheet.max_row}"
    assert sheet[1][0].font.bold is True
    assert sheet.column_dimensions["B"].width > 0
    assert sheet.column_dimensions["G"].width > 0
    if sheet.max_row > 1:
        assert sheet[2][1].alignment.wrap_text is True
        assert sheet[2][6].alignment.wrap_text is True
        assert sheet[2][8].hyperlink is not None
        assert sheet[2][8].hyperlink.target == sheet[2][8].value


def test_export_empty_view_still_returns_valid_header(client, monkeypatch):
    monkeypatch.setattr(app, "offline_paginate_tickets", lambda: [])
    response = client.get("/queue/export.xlsx?mode=normal&photo_video_only=1&hide_reviewed_tags=1&missing_tags=1&workflow_tab=main")
    workbook = load_workbook(io.BytesIO(response.data))
    sheet = workbook["Current View"]
    assert response.status_code == 200
    assert sheet.max_row == 1
    assert [cell.value for cell in sheet[1]][:2] == ["Ticket #", "Subject"]
    assert sheet.freeze_panes == "A2"


def test_export_control_preserves_canonical_state(client):
    response = client.get("/queue?mode=closed&photo_video_only=0&missing_tags=0&workflow_tab=resolved")
    assert response.status_code == 200
    match = re.search(rb'href="(/queue/export\.xlsx\?[^\"]+)"', response.data)
    assert match
    href = match.group(1).decode()
    assert "mode=closed" in href
    assert "photo_video_only=0" in href
    assert "missing_tags=0" in href
    assert "workflow_tab=resolved" in href
