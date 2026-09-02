import io

from openpyxl import load_workbook

import app


def _ticket(ticket_id=950001, **changes):
    ticket = {
        "id": ticket_id,
        "subject": "Photo/Video Request Customer 445333",
        "status": 2,
        "type": "Guest Callback/Follow-Up",
        "custom_fields": {"cf_follow_up_group": "Service"},
        "group_id": 154000437139,
        "tags": ["PHOTOS"],
        "priority": 2,
        "created_at": "2026-09-01T12:00:00Z",
        "updated_at": "2026-09-01T12:00:00Z",
    }
    ticket.update(changes)
    return ticket


def test_main_queue_accepts_all_required_statuses():
    for status in app.MAIN_QUEUE_STATUSES:
        assert app.is_main_queue_ticket(_ticket(status=status))
    assert app.main_queue_triage_reasons(_ticket(status=4)) == [app.TRIAGE_REASON_STATUS]


def test_main_queue_type_group_and_tag_rules():
    assert app.main_queue_triage_reasons(_ticket(type="Other")) == [app.TRIAGE_REASON_TYPE]
    assert app.main_queue_triage_reasons(_ticket(group_id=154000472479)) == [app.TRIAGE_REASON_GROUP]
    assert app.main_queue_triage_reasons(_ticket(tags=["other"])) == [app.TRIAGE_REASON_TAG]


def test_subject_is_not_a_main_queue_gate():
    for subject in ("Photo request", "Photo/video", "Unrelated subject text", ""):
        assert app.is_main_queue_ticket(_ticket(subject=subject))
        assert app.main_queue_triage_reasons(_ticket(subject=subject)) == []


def test_native_service_group_is_authoritative():
    assert app.is_main_queue_ticket(_ticket(custom_fields={"cf_follow_up_group": "Retail"}))
    assert app.is_main_queue_ticket(_ticket(custom_fields={}))
    assert app.main_queue_triage_reasons(_ticket(group_id="154000437139")) == []
    assert app.main_queue_triage_reasons(_ticket(group_id=154000472479)) == [app.TRIAGE_REASON_GROUP]
    assert app.main_queue_triage_reasons(_ticket(group_id=None)) == [app.TRIAGE_REASON_GROUP]


def test_ticket_444185_like_custom_service_group_is_triage():
    ticket = _ticket(
        ticket_id=444185,
        group_id=154000472479,
        custom_fields={"cf_follow_up_group": "Service"},
    )
    assert not app.is_main_queue_ticket(ticket)
    assert app.main_queue_triage_reasons(ticket) == [app.TRIAGE_REASON_GROUP]


def test_each_dashboard_tag_is_case_and_whitespace_tolerant():
    for tag in app.MAIN_QUEUE_PHOTO_VIDEO_TAGS:
        assert app.has_main_queue_photo_video_tag(_ticket(tags=[f"  {tag.upper()}  "]))


def test_subject_matcher_accepts_request_variations_and_rejects_loose_words():
    for subject in (
        "Photo/Video Request #444114",
        "Photo/video request Kyle 3014374935 #445333",
        "PHOTO VIDEO REQUEST AIMEE 2084363220 #444106",
        "Video/Photo Request #445325",
        "Video / Photo Request",
        "Photo / Video Request",
    ):
        assert app.subject_matches_main_queue_photo_video_request(_ticket(subject=subject))
    for subject in ("Photo attached", "Video available", "Request for a photo"):
        assert not app.subject_matches_main_queue_photo_video_request(_ticket(subject=subject))


def test_triage_reasons_are_complete_and_deterministic():
    ticket = _ticket(status=4, type="Other", group_id=154000472479, tags=[], subject="Photo attached")
    assert app.main_queue_triage_reasons(ticket) == [
        app.TRIAGE_REASON_STATUS,
        app.TRIAGE_REASON_TYPE,
        app.TRIAGE_REASON_GROUP,
        app.TRIAGE_REASON_TAG,
    ]


def test_queue_partition_and_review_identity(monkeypatch):
    main = _ticket(950010)
    triage = _ticket(950011, tags=[])
    config = dict(app.DEFAULT_FILTERS, photo_video_only=False, hide_reviewed_tags=False, workflow_tab="main")
    monkeypatch.setattr(app, "load_review_rows", lambda: {950010: {"review_result": "Resolved"}, 950011: {"review_result": "Resolved"}})
    monkeypatch.setattr(app, "last_opened_ticket_id", lambda: None)
    monkeypatch.setattr(app, "updated_since_review", lambda ticket, state: False)
    main_rows, _, counts, *_ = app.build_current_queue_view([main, triage], dict(config, queue_scope="main"))
    triage_rows, _, _, *_ = app.build_current_queue_view([main, triage], dict(config, queue_scope="triage"))
    assert counts == {"main": 1, "triage": 1}
    assert [row["id"] for row in main_rows] == [950010]
    assert [row["id"] for row in triage_rows] == [950011]
    assert main_rows[0]["result"] == triage_rows[0]["result"] == "Resolved"
    assert triage_rows[0]["triage_reasons"] == [app.TRIAGE_REASON_TAG]


def test_queue_scope_and_workflow_labels_are_distinct(client):
    response = client.get("/queue")
    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert 'aria-label="Queue scope"' in html
    assert ">Main Queue <span" in html
    assert ">Needs Triage <span" in html
    assert ">To Review <span" in html
    for label in ("Supervisor Review", "Follow-Up", "Resolved", "No Action"):
        assert f">{label} <span" in html
    assert app.WORKFLOW_LABELS["main"] == "To Review"


def test_queue_scope_defaults_main_and_closed_mode_does_not_split(client):
    assert app.filters_from_args({}).get("queue_scope", "main") == "main"
    assert app.filters_from_args({"queue_scope": "unexpected"}).get("queue_scope", "main") == "main"
    response = client.get("/queue?mode=closed")
    assert response.status_code == 200
    assert b'aria-label="Queue scope"' not in response.data


def test_triage_export_has_why_here_and_respects_scope(client, monkeypatch):
    triage = _ticket(950020, tags=[])
    monkeypatch.setattr(app, "get_ticket_pool", lambda: ([triage], None))
    response = client.get("/queue/export.xlsx?queue_scope=triage&photo_video_only=0&hide_reviewed_tags=0")
    assert response.status_code == 200
    sheet = load_workbook(io.BytesIO(response.data)).active
    headers = [cell.value for cell in sheet[1]]
    assert "Why here?" in headers
    assert sheet.cell(2, headers.index("Why here?") + 1).value == app.TRIAGE_REASON_TAG


def test_main_export_omits_triage_reason_column(client, monkeypatch):
    main = _ticket(950021)
    monkeypatch.setattr(app, "get_ticket_pool", lambda: ([main], None))
    response = client.get("/queue/export.xlsx?queue_scope=main&photo_video_only=0&hide_reviewed_tags=0")
    sheet = load_workbook(io.BytesIO(response.data)).active
    assert "Why here?" not in [cell.value for cell in sheet[1]]
