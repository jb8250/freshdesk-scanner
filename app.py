"""Freshdesk Review Queue Dashboard — read-only scanner + local review workflow.

Routes:
  GET  /queue               dashboard + local-only filter controls (render-only)
  POST /queue/api/refresh   start one explicit finite background Freshdesk refresh
  GET  /queue/api/refresh/status
                            local-only progress snapshot for that refresh
  POST /queue/api/refresh/cancel
                            cooperatively cancel the active queue refresh
  POST /queue/api/review    save a local review result (form POST, CSRF-protected)
  POST /queue/api/opened    record that a ticket link was opened (JSON, CSRF-protected)
  GET  /closed              closed-ticket housekeeping (render-only, cache only)
  POST /closed/api/refresh  explicit manual retrieval for the closed cache
  POST /closed/api/review   save a local closed-ticket review result
  POST /closed/api/opened   record that a closed ticket link was opened

Data sources (strictly isolated):
  live    POST /queue/api/refresh -> background GET /api/v2/tickets on the Freshdesk account
          (read-only list endpoint; write methods POST/PUT/PATCH/DELETE never
          used). Results are written to the LIVE queue cache
          (cache/queue_live_tickets.json).
  offline FRESHDESK_OFFLINE=1 -> local fixture pages (fixtures/fixtures.json),
          no network, no API key.

Network discipline (mandatory):
   * GET /queue and GET /closed are RENDER ONLY. They must NEVER trigger a
     Freshdesk request — not on browser reload or for a missing/stale cache.
     Freshness/TTL is shown as information only; the LIVE scheduler performs
     its separate normal refresh only after its 30-minute deadline.

   * Queue Freshdesk retrieval happens through the normal finite worker after
     an explicit, CSRF-protected POST /queue/api/refresh or the LIVE process's
     30-minute automatic scheduler. Closed retrieval remains POST
     /closed/api/refresh. A bounded pagination/rate-limit sequence within one
     refresh is permitted.

  * Freshdesk is GET-only for data: no POST/PUT/PATCH/DELETE to the API.

Cache isolation:
  * LIVE queue cache  -> cache/queue_live_tickets.json  (written after successful Refresh only)
  * OFFLINE fixtures  -> fixtures/fixtures.json          (loaded directly)
  * LIVE closed cache -> cache/closed_tickets.json       (written by Refresh only)
  The live queue cache is never read from the fixtures file and the fixtures
  are never read from the live cache; each is addressed at a distinct path.

Local review state (never sent to Freshdesk): SQLite at data/review_state.sqlite3
(override with REVIEW_DB_PATH). Contains per-ticket review result, first/last
opened timestamps, last review change, and the ticket updated_at snapshot taken
when a reviewed state was assigned (drives the Updated Since Review flag).

Offline mode is fail-closed: it never calls the network and never reads the API
key. If the fixture data is missing or malformed, /queue renders an error page
instead of falling back to live access.

The API key is never loaded at import time. Only load_api_key() touches the key
file, and only the live retrieval path calls it.
"""
import hmac
import json
import math
import os
import re
import secrets
import sqlite3
import tempfile
import threading
import io
import time
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from math import ceil
from typing import Callable, Optional
from zoneinfo import ZoneInfo
from urllib.parse import urlencode

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from flask import (Flask, jsonify, redirect, render_template_string, request,
                   send_file, session)

import review_backups

# Live Closed dashboard support (Prompt 24): separate cache, atomic writes and
# the single-slot refresh job. Importing it here is safe — closed_live imports
# closed_retriever lazily, so there is no import cycle.
import closed_live
import queue_live
from auto_refresh import AUTO_REFRESH_INTERVAL_SECONDS, AutoRefreshScheduler
from queue_merge import merge_queue_tickets
from queue_retention import apply_queue_retention

app = Flask(__name__, static_folder=None)
# Per-process random key: used only for the loopback CSRF token and flash
# messages. Changes every restart, which is correct for a local tool.
app.secret_key = secrets.token_hex(32)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def is_offline() -> bool:
    """Explicit offline mode flag. Read from the environment on every call so
    tests (and operators) can flip it without a restart. There is no
    auto-detection: if FRESHDESK_OFFLINE is not set, the app is in live mode.
    """
    return os.environ.get("FRESHDESK_OFFLINE", "").strip().lower() in ("1", "true", "yes")


# Freshdesk queue scanner config
FRESHDESK_DOMAIN = "broadriverretail-help.freshdesk.com"
FRESHDESK_KEY_FILE = os.path.expanduser("~/.config/furtouch/freshdesk_api_key")


def _safe_float_env(name, default, minimum):
    try:
        value = float(os.environ.get(name, str(default)))
        if not math.isfinite(value):
            raise ValueError
        return max(minimum, value)
    except (TypeError, ValueError):
        return default


def _safe_int_env(name, default, minimum):
    try:
        value = int(os.environ.get(name, str(default)))
        return max(minimum, value)
    except (TypeError, ValueError):
        return default


FRESHDESK_MIN_REQUEST_INTERVAL_SECONDS = _safe_float_env(
    "FRESHDESK_MIN_REQUEST_INTERVAL_SECONDS", 6, 1
)
FRESHDESK_MIN_REMAINING = _safe_int_env("FRESHDESK_MIN_REMAINING", 20, 0)
FRESHDESK_MAX_RETRIES = min(_safe_int_env("FRESHDESK_MAX_RETRIES", 2, 0), 3)
QUEUE_RETRIEVAL_LOCK = threading.Lock()  # process-local guard for this Flask process

# Never populated at import time. Live mode only.
FRESHDESK_API_KEY = ""

# Offline fixtures: JSON file containing {"pages": [[ticket, ...], ...]}.
FIXTURES_FILE = os.environ.get(
    "FRESHDESK_FIXTURES",
    os.path.join(BASE_DIR, "fixtures", "fixtures.json"),
)

CACHE_DIR = os.path.join(BASE_DIR, "cache")
os.makedirs(CACHE_DIR, exist_ok=True)

# LIVE queue cache — written ONLY after an explicit successful queue refresh
# retrieval. GET /queue reads only this file (and never the fixtures file),
# so offline data can never satisfy a live cache read and vice versa.
LIVE_QUEUE_CACHE_FILE = os.path.join(CACHE_DIR, "queue_live_tickets.json")
# Historical queue cache name kept for backward compatibility of read paths;
# the live retrieval writes to LIVE_QUEUE_CACHE_FILE.
CACHE_FILE = LIVE_QUEUE_CACHE_FILE
CACHE_TTL_SECONDS = 30 * 60  # informational only in live mode (never auto-fetches)

# Queue-cache envelope v2 is deliberately separate from ticket data.  Future
# phases may use the successful refresh start timestamp as their cursor, but
# Phase 3A continues to replace the entire cache after every successful refresh.
QUEUE_CACHE_SCHEMA_VERSION = 2
ROLLING_RETENTION_DAYS = 60
QUEUE_CACHE_REFRESH_MODES = frozenset({"legacy", "baseline", "reconcile", "incremental", "full_rebuild"})
INCREMENTAL_CURSOR_OVERLAP_SECONDS = 120

UPDATED_SINCE_DAYS = 60  # ~2 months

# Scanner keyword set — matches Chrome extension logic (word-boundary regex).
# Used by the default Review Scope to qualify photo/video ticket subjects or tags.
KEYWORDS = [
    "photo", "photos", "picture", "pictures",
    "pic", "pics", "video", "videos", "vid", "vids",
]
# Underscore is treated as a separator alongside punctuation/whitespace so tags
# such as ``Video_photo request`` qualify without allowing loose substrings.
KEYWORD_RE = re.compile(r"(?<![^\W_])(" + "|".join(KEYWORDS) + r")(?![^\W_])", re.IGNORECASE)

MAIN_QUEUE_STATUSES = frozenset({2, 6, 7, 8})
MAIN_QUEUE_TYPE = "guest callback/follow-up"
MAIN_QUEUE_GROUP_ID = 154000437139
MAIN_QUEUE_PHOTO_VIDEO_TAGS = frozenset({
    "photo/video request",
    "photo request",
    "video/ photos",
    "photos",
    "video/ photo request",
    "video/ photo",
    "product issue video request",
})
PHOTO_VIDEO_REQUEST_SUBJECT_RE = re.compile(
    r"\b(?:photo\s*/?\s*video|video\s*/?\s*photo)\s+request\b", re.IGNORECASE
)
TRIAGE_REASON_STATUS = "Status not in Main Queue"
TRIAGE_REASON_TYPE = "Wrong Type"
TRIAGE_REASON_GROUP = "Wrong Group"
TRIAGE_REASON_TAG = "Missing photo/video tag"

# Reviewed/closed Freshdesk tags (human-applied). Any ONE of these removes a
# ticket from the DEFAULT review queue. Comparison is case-insensitive and
# leading/trailing-whitespace-insensitive; stored ticket tags are never
# mutated. This set is intentionally limited to these six values.
REVIEWED_EXCLUSION_TAGS = frozenset({
    "parts needed",
    "exchange",
    "no service needed",
    "closed",
    "schedule service",
    "delivery special needed",
})


class OfflineDataError(Exception):
    """Raised when offline mode cannot load valid fixture data. The app must
    fail closed on this — never fall back to the live API."""


_auto_refresh_scheduler = None
_auto_refresh_scheduler_lock = threading.Lock()


def _automatic_queue_refresh():
    """Start the normal queue refresh when the monotonic scheduler is due."""
    if is_offline():
        return "offline"
    if queue_live.JOB.status()["running"]:
        return "skipped_busy"
    started, message = _start_normal_queue_refresh()
    if started:
        return "started"
    return "skipped_busy" if queue_live.JOB.status()["running"] else "failed"


def start_auto_refresh_scheduler():
    """Start the one live-process automatic-refresh scheduler, if needed."""
    global _auto_refresh_scheduler
    if is_offline():
        return False
    with _auto_refresh_scheduler_lock:
        if _auto_refresh_scheduler is None:
            _auto_refresh_scheduler = AutoRefreshScheduler(_automatic_queue_refresh)
        return _auto_refresh_scheduler.start()


def initialize_live_auto_refresh():
    """Production startup hook; safe for direct-app launch and testable without Flask CLI."""
    return False if is_offline() else start_auto_refresh_scheduler()


def reset_auto_refresh_countdown():
    """Rearm an already-running live scheduler after an accepted user job."""
    scheduler = _auto_refresh_scheduler
    if scheduler is not None and not is_offline():
        scheduler.reset()


def auto_refresh_status():
    """Return local scheduler metadata only; never load credentials or data."""
    scheduler = _auto_refresh_scheduler
    if is_offline() or scheduler is None:
        return {"enabled": False, "interval_seconds": AUTO_REFRESH_INTERVAL_SECONDS,
                "seconds_until_next": None, "last_attempt_at": None, "last_result": None}
    return scheduler.status()


def load_api_key() -> str:
    """Load the Freshdesk API key (live mode only). Reads FRESHDESK_API_KEY env
    var first, then falls back to the chmod-600 key file. The key file is never
    touched while offline mode is active. Returns "" when no key is available.
    """
    global FRESHDESK_API_KEY
    if FRESHDESK_API_KEY:
        return FRESHDESK_API_KEY
    key = os.environ.get("FRESHDESK_API_KEY", "").strip()
    if not key and not is_offline() and os.path.exists(FRESHDESK_KEY_FILE):
        with open(FRESHDESK_KEY_FILE, "r") as fh:
            key = fh.read().strip()
    FRESHDESK_API_KEY = key
    return key


def fd_auth():
    return (load_api_key(), "X")


def keyword_filter_hits(text):
    return bool(KEYWORD_RE.search(text or ""))


# ---------------------------------------------------------------------------
# Default Review Scope (Phase 1): two visible, independent scope layers.
#
# Layer 1 — photo/video scope: the ticket SUBJECT or any valid Freshdesk TAG
# must contain one of the recognized photo/video keywords (case-insensitive,
# word-boundary aware). Description/body/notes/conversations/attachments/
# requester data are never consulted for this rule.
#
# Layer 2 — reviewed/closed tag exclusions: a ticket carrying ANY one of the
# six REVIEWED_EXCLUSION_TAGS is hidden from the default review queue.
# Comparison normalizes case and surrounding whitespace for comparison only;
# stored Freshdesk tags are never modified.
#
# Manual queue filters are separate and remain opt-in (neutral by default).
# ---------------------------------------------------------------------------


def text_matches_photo_video(text):
    """True when text contains a recognized photo/video keyword.

    Missing, None, and non-string text fail safely. The shared regex remains
    case-insensitive and word-boundary aware.
    """
    return isinstance(text, str) and bool(KEYWORD_RE.search(text))


def subject_matches_photo_video(ticket):
    """Backward-compatible subject-only photo/video matcher."""
    return isinstance(ticket, dict) and text_matches_photo_video(ticket.get("subject"))


def normalized_queue_value(value):
    return " ".join(value.strip().casefold().split()) if isinstance(value, str) else ""


def main_queue_group_id(ticket):
    value = ticket.get("group_id") if isinstance(ticket, dict) else None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def main_queue_status(ticket):
    value = ticket.get("status") if isinstance(ticket, dict) else None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def has_main_queue_photo_video_tag(ticket):
    tags = ticket.get("tags") if isinstance(ticket, dict) else None
    return isinstance(tags, list) and any(
        normalized_queue_value(tag) in MAIN_QUEUE_PHOTO_VIDEO_TAGS for tag in tags
    )


def subject_matches_main_queue_photo_video_request(ticket):
    subject = ticket.get("subject") if isinstance(ticket, dict) else None
    return isinstance(subject, str) and bool(PHOTO_VIDEO_REQUEST_SUBJECT_RE.search(subject))


def main_queue_triage_reasons(ticket):
    """Return stable, complete Main Queue rule failures for one candidate."""
    reasons = []
    if main_queue_status(ticket) not in MAIN_QUEUE_STATUSES:
        reasons.append(TRIAGE_REASON_STATUS)
    if normalized_queue_value(ticket.get("type") if isinstance(ticket, dict) else None) != MAIN_QUEUE_TYPE:
        reasons.append(TRIAGE_REASON_TYPE)
    if main_queue_group_id(ticket) != MAIN_QUEUE_GROUP_ID:
        reasons.append(TRIAGE_REASON_GROUP)
    if not has_main_queue_photo_video_tag(ticket):
        reasons.append(TRIAGE_REASON_TAG)
    return reasons


def is_main_queue_ticket(ticket):
    return not main_queue_triage_reasons(ticket)


def ticket_matches_photo_video(ticket):
    """True when a ticket's subject or any valid Freshdesk tag matches.

    Tags must be a list; missing, malformed, and non-string tag values are
    ignored safely without mutating the source ticket.
    """
    if not isinstance(ticket, dict):
        return False
    if text_matches_photo_video(ticket.get("subject")):
        return True
    tags = ticket.get("tags")
    return isinstance(tags, list) and any(text_matches_photo_video(tag) for tag in tags)


def normalized_ticket_tags(ticket):
    """Casefolded + stripped view of a ticket's tag strings, for comparison
    only. Never mutates the source ticket. Non-list/missing tags and non-string
    tags are ignored safely."""
    raw = ticket.get("tags")
    if not isinstance(raw, list):
        return set()
    normalized = set()
    for tag in raw:
        if not isinstance(tag, str):
            continue
        norm = tag.strip().casefold()
        if norm:
            normalized.add(norm)
    return normalized


def has_reviewed_exclusion_tag(ticket):
    """True when the ticket carries ANY of the six reviewed/closed exclusion
    tags (case-insensitive, whitespace-insensitive comparison only)."""
    return bool(normalized_ticket_tags(ticket) & set(REVIEWED_EXCLUSION_TAGS))


def passes_review_scope(ticket, config):
    """Default working review scope layer.

    Runs BEFORE the opt-in manual filters. Both scope fields default ON;
    missing keys also default ON so a hand-built config can never silently
    widen the default review queue. With photo_video_only OFF the subject rule
    imposes no restriction; with hide_reviewed_tags OFF the exclusion tags
    impose no restriction. Both OFF means every cached ticket passes this
    layer (subject only to ticket-ID deduplication later in the pipeline).
    """
    photo_video_only = config.get("photo_video_only", True)
    hide_reviewed_tags = config.get("hide_reviewed_tags", True)
    if photo_video_only and not ticket_matches_photo_video(ticket):
        return False
    if hide_reviewed_tags and has_reviewed_exclusion_tag(ticket):
        return False
    return True


# Known status values on this account (from live ticket data):
#   2 = Customer responded  (needs review)
#   5 = Closed               (exclude)
#   6 = Waiting on customer  (optional include)
SCAN_STATUSES = [2, 6]  # Customer responded + Waiting on customer

STATUS_LABELS = {2: "Customer responded", 3: "Pending", 4: "Resolved", 5: "Closed", 6: "Waiting on customer", 1: "Open"}
PRIORITY_LABELS = {1: "Low", 2: "Medium", 3: "High", 4: "Urgent"}

# Freshdesk Filter Tickets API v2 contract, documented in
# docs/closed_housekeeping_api_contract.md. This closed-housekeeping foundation
# deliberately has no live HTTP adapter: all retrieval is fake/injectable.
CLOSED_STATUS = 5


def status_label(value) -> str:
    """Display a single ticket status value as text.

    Only the integer CLOSED_STATUS (5) is labelled "Closed"; any other value —
    another known integer status, an unknown number, a string, or a missing
    value — renders its real label or a neutral fallback. It can never be
    silently labelled "Closed", so invalid rows fail visibly instead of being
    mislabelled. The integer status (5) remains the internal/API filter value;
    this is presentation-only.
    """
    if value == CLOSED_STATUS and isinstance(value, int) and not isinstance(value, bool):
        return STATUS_LABELS[CLOSED_STATUS]  # "Closed"
    if isinstance(value, int) and not isinstance(value, bool) and value in STATUS_LABELS:
        return STATUS_LABELS[value]
    return "Unknown"


SEARCH_PAGE_SIZE = 30
SEARCH_MAX_PAGE = 10
SEARCH_MAX_RESULTS = SEARCH_PAGE_SIZE * SEARCH_MAX_PAGE
CLOSED_MAX_SPLIT_DEPTH = 20
CLOSED_MAX_WINDOWS = 512
CLOSED_DEFAULT_DAYS = 60
CLOSED_MIN_DAYS = 1
CLOSED_MAX_DAYS = 3650

# ---------------------------------------------------------------------------
# Dashboard filter configuration (URL-backed, documented defaults)
# ---------------------------------------------------------------------------

# Queue filters are deliberately neutral by default. Freshdesk retrieval and
# local filtering are separate operations: a refresh fills the cache, and the
# operator opts into any local restriction afterward. Days is a retrieval
# setting, not a local result filter.
#
# The two Review Scope fields (photo_video_only / hide_reviewed_tags) are
# SEPARATE from the manual filters below. They define the visible default
# working review queue and default ON. The manual filters default OFF and stay
# independent; the explicit "Show All Cached Tickets" control turns both scope
# fields plus every manual filter off to display the complete cache.
DEFAULT_FILTERS = {
    "mode": "normal",
    "photo_video_only": True,    # default Review Scope: subject matches photo/video keyword
    "hide_reviewed_tags": True,  # default Review Scope: no reviewed/closed exclusions
    "overdue": False,            # opt-in: due_by is a valid timestamp earlier than now
    "responded": False,          # opt-in: status == 2 (Customer responded)
    "waiting": False,            # opt-in: status == 6 (Waiting on customer)
    "missing_tags": False,       # opt-in: tags absent or empty
    "days": 60,                  # Freshdesk retrieval window (1-365)
    "review_view": "all",        # legacy compatibility; workflow_tab is canonical
    "workflow_tab": "main",      # default workflow tab
}
DAYS_MIN, DAYS_MAX, DAYS_DEFAULT = 1, 365, 60
REVIEW_VIEWS = ("active", "completed", "all")
WORKFLOW_TABS = ("main", "supervisor", "followup", "resolved", "no_action")
WORKFLOW_LABELS = {
    "main": "To Review", "supervisor": "Supervisor Review", "followup": "Follow-Up",
    "resolved": "Resolved", "no_action": "No Action",
}

# Local review results (stored in SQLite only — never sent to Freshdesk).
REVIEW_STATES = [
    "Unreviewed",
    "Opened / In Review",
    "Needs Supervisor Review",
    "Resolved",
    "Not Applicable to Me",
    "No Action Needed",
    "Needs Follow-Up",
]
# States that snapshot the ticket's updated_at at review time. A later ticket
# update compared against that snapshot produces the "UPDATED SINCE REVIEW"
# flag, and such tickets are treated as Active again.
REVIEWED_STATES = {"Resolved", "Not Applicable to Me", "No Action Needed", "Needs Follow-Up", "Needs Supervisor Review"}
ACTIVE_STATES = {"Unreviewed", "Opened / In Review", "Needs Follow-Up", "Needs Supervisor Review"}
COMPLETED_STATES = {"Resolved", "Not Applicable to Me", "No Action Needed"}


def parse_workflow_tab(value):
    return value if value in WORKFLOW_TABS else "main"


def workflow_destination(state, updated=False):
    """Return the local workflow destination; updates always route to Main."""
    if updated:
        return "main"
    return {
        "Unreviewed": "main", "Opened / In Review": "main",
        "Needs Supervisor Review": "supervisor", "Needs Follow-Up": "followup",
        "Resolved": "resolved", "No Action Needed": "no_action",
        "Not Applicable to Me": "no_action",
    }.get(state, "main")


def workflow_tab_includes(state_row, updated_flag, tab):
    state = state_row.get("review_result", "Unreviewed") if state_row else "Unreviewed"
    return workflow_destination(state, updated_flag) == tab


def human_age(seconds):
    """Compact, deterministic age label for cache metadata."""
    try:
        seconds = float(seconds)
    except (TypeError, ValueError):
        return "Unknown"
    if seconds < 0 or not math.isfinite(seconds):
        return "Unknown"
    if seconds < 60:
        return "Just now"
    minutes = int(seconds // 60)
    if minutes < 60:
        return f"{minutes}m ago"
    hours = minutes // 60
    if hours < 24:
        return f"{hours}h ago"
    return f"{hours // 24}d ago"

# ---------------------------------------------------------------------------
# Time helpers (monkeypatchable in tests)
# ---------------------------------------------------------------------------


def now_utc():
    return datetime.now(timezone.utc)


def iso_now():
    return now_utc().isoformat()


def parse_dt(value):
    """Parse an offset-aware ISO-8601 timestamp with optional Z suffix.

    Timezone-less values are rejected so callers can fail closed instead of
    comparing a naive datetime with the dashboard's UTC-aware timestamps.
    """
    if not value or not isinstance(value, str):
        return None
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except Exception:
        return None
    return parsed if parsed.tzinfo is not None else None


# ---------------------------------------------------------------------------
# URL parameter parsing (safe fallbacks)
# ---------------------------------------------------------------------------


def _last_value(args, key):
    """Last occurrence wins for repeated query values."""
    try:
        vals = args.getlist(key)
    except (KeyError, AttributeError):
        return None
    return vals[-1] if vals else None


def parse_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    low = str(value).strip().lower()
    if low in ("1", "true", "yes", "on"):
        return True
    if low in ("0", "false", "no", "off"):
        return False
    return default  # invalid -> documented default


def parse_days(value):
    if value is None:
        return DAYS_DEFAULT
    s = str(value).strip()
    if not s or not s.isdigit():
        return DAYS_DEFAULT
    n = int(s)
    if not (DAYS_MIN <= n <= DAYS_MAX):
        return DAYS_DEFAULT
    return n


def parse_review_view(value, default=None):
    if value in REVIEW_VIEWS:
        return value
    if default in REVIEW_VIEWS:
        return default
    return DEFAULT_FILTERS["review_view"]  # queue invalid/missing -> neutral All


def parse_queue_mode(value):
    """Return a stable queue mode; missing and invalid values stay normal."""
    return "closed" if value == "closed" else "normal"


def parse_queue_scope(value):
    return "triage" if value == "triage" else "main"


def filters_from_args(args, submitted=False):
    """Build a canonical local queue configuration from request values.

    The removed overdue/responded/waiting controls remain accepted as inert URL
    parameters so old bookmarks cannot restrict or break the unified queue.
    """
    mode = parse_queue_mode(_last_value(args, "mode"))
    checkbox_default = False if submitted else None

    def checkbox(name, default=None):
        value = _last_value(args, name)
        fallback = checkbox_default if submitted else (DEFAULT_FILTERS[name] if default is None else default)
        return parse_bool(value, fallback)

    cfg = {
        "mode": mode,
        "photo_video_only": checkbox("photo_video_only", True if mode == "closed" else None),
        "hide_reviewed_tags": checkbox("hide_reviewed_tags", False if mode == "closed" else None),
        # Retain legacy values for canonical bookmarked URLs. Queue filtering
        # deliberately ignores them; their controls no longer exist.
        "overdue": checkbox("overdue"),
        "responded": checkbox("responded"),
        "waiting": checkbox("waiting"),
        "missing_tags": checkbox("missing_tags", True if mode == "closed" else None),
        "days": parse_days(_last_value(args, "days")),
        "review_view": parse_review_view(_last_value(args, "review_view")),
    }
    workflow_tab = _last_value(args, "workflow_tab")
    cfg["workflow_tab"] = parse_workflow_tab(workflow_tab) if workflow_tab is not None else DEFAULT_FILTERS["workflow_tab"]
    if parse_queue_scope(_last_value(args, "queue_scope")) == "triage":
        cfg["queue_scope"] = "triage"
    # Closed mode carries the immediately preceding Normal Review workspace in
    # private URL parameters. Keep it absent for a direct Closed entry: there
    # is no saved Normal workspace to synthesize in that case.
    if mode == "closed" and any(_last_value(args, f"normal_{name}") is not None for name in NORMAL_RETURN_FIELDS):
        cfg["normal_return"] = normal_return_from_args(args)
    return cfg


NORMAL_RETURN_FIELDS = (
    "photo_video_only", "hide_reviewed_tags", "missing_tags", "days",
    "review_view", "workflow_tab", "queue_scope",
)


def normal_return_from_args(args):
    """Canonicalize the private Normal workspace preserved in Closed URLs."""
    return {
        "photo_video_only": parse_bool(_last_value(args, "normal_photo_video_only"), DEFAULT_FILTERS["photo_video_only"]),
        "hide_reviewed_tags": parse_bool(_last_value(args, "normal_hide_reviewed_tags"), DEFAULT_FILTERS["hide_reviewed_tags"]),
        "missing_tags": parse_bool(_last_value(args, "normal_missing_tags"), DEFAULT_FILTERS["missing_tags"]),
        "days": parse_days(_last_value(args, "normal_days")),
        "review_view": parse_review_view(_last_value(args, "normal_review_view")),
        "workflow_tab": parse_workflow_tab(_last_value(args, "normal_workflow_tab")),
        "queue_scope": parse_queue_scope(_last_value(args, "normal_queue_scope")),
    }


def filter_query_string(config):
    """Canonical queue query string, retaining Normal return state only in Closed mode."""
    mode = parse_queue_mode(config.get("mode"))
    params = {
        "mode": mode,
        "photo_video_only": "1" if config.get("photo_video_only", True) else "0",
        "hide_reviewed_tags": "1" if config.get("hide_reviewed_tags", True) else "0",
        "overdue": "1" if config.get("overdue") else "0",
        "responded": "1" if config.get("responded") else "0",
        "waiting": "1" if config.get("waiting") else "0",
        "missing_tags": "1" if config.get("missing_tags") else "0",
        "days": str(config.get("days", DAYS_DEFAULT)),
        "review_view": config.get("review_view", "all"),
        "workflow_tab": parse_workflow_tab(config.get("workflow_tab", "main")),
    }
    if parse_queue_scope(config.get("queue_scope", "main")) == "triage":
        params["queue_scope"] = "triage"
    normal_return = config.get("normal_return")
    if mode == "closed" and isinstance(normal_return, dict):
        for name in NORMAL_RETURN_FIELDS:
            value = normal_return.get(name)
            if name in ("photo_video_only", "hide_reviewed_tags", "missing_tags"):
                params[f"normal_{name}"] = "1" if value else "0"
            elif name == "days":
                params[f"normal_{name}"] = str(parse_days(value))
            elif name == "review_view":
                params[f"normal_{name}"] = parse_review_view(value)
            elif name == "queue_scope":
                params[f"normal_{name}"] = parse_queue_scope(value)
            else:
                params[f"normal_{name}"] = parse_workflow_tab(value)
    return urlencode(params)


_VIEW_LABEL = {"active": "Active", "completed": "Completed", "all": "All"}


def _last_refresh_display(cache_age, offline):
    if offline or cache_age is None:
        return "Never"
    return human_age(cache_age)


def _cache_coverage_display(cached_days, selected_days):
    if not isinstance(cached_days, int):
        return "Unknown"
    label = f"Last {cached_days} day{'s' if cached_days != 1 else ''} (covers the last {cached_days} day{'s' if cached_days != 1 else ''})"
    if isinstance(selected_days, int) and selected_days > cached_days:
        return f"{label} (selected {selected_days})"
    return label


def filter_summary_text(config):
    """Human-readable summary of *local* queue restrictions only.

    Days is intentionally omitted because it controls the Freshdesk retrieval
    window rather than filtering the already-cached rows. The Review Scope
    fields are shown first because they define the default working queue;
    manual filters and the review view are shown after them. With every scope
    field and manual filter off and Review View=All, the queue is a complete
    view of the cache.
    """
    segments = []
    if config.get("photo_video_only"):
        segments.append("Photo/video subjects only")
    if config.get("hide_reviewed_tags"):
        segments.append("No reviewed/closed tags")
    if config.get("overdue"):
        segments.append("Overdue")
    if config.get("responded"):
        segments.append("Customer Responded")
    if config.get("waiting"):
        segments.append("Waiting on Customer")
    if config.get("missing_tags"):
        segments.append("Missing Tags")
    view = config.get("review_view", "all")
    if not segments and view == "all":
        return "Showing: All cached tickets"
    if not segments:
        segments.append("All ticket conditions")
    if view != "all":
        segments.append(f"{_VIEW_LABEL.get(view, 'All')} review view")
    return f"Showing: {' + '.join(segments)}"



# ---------------------------------------------------------------------------
# Category logic (section 6 of the spec)
# ---------------------------------------------------------------------------


def is_overdue(t):
    """Overdue: due_by exists, is a valid timestamp, and is earlier than now.
    Missing or malformed due_by is NOT overdue (fail-closed)."""
    dt = parse_dt(t.get("due_by"))
    return dt is not None and dt < now_utc()


def is_customer_responded(t):
    return t.get("status") == 2


def is_waiting_on_customer(t):
    return t.get("status") == 6


def has_missing_tags(t):
    tags = t.get("tags")
    return not tags or not isinstance(tags, list) or len(tags) == 0


def category_matches(t, category):
    if category == "overdue":
        return is_overdue(t)
    if category == "responded":
        return is_customer_responded(t)
    if category == "waiting":
        return is_waiting_on_customer(t)
    return False


def has_primary_filter(config):
    """Return whether any primary condition/status checkbox is selected.

    This is informational only. A false result now means "no primary
    restriction", not "show no rows".
    """
    return config["overdue"] or config["responded"] or config["waiting"]


def matches_status_group(t, config):
    """OR within the status group (Customer Responded / Waiting on Customer).
    The two are mutually exclusive Freshdesk statuses: selecting exactly one
    shows only that status, selecting both shows either. When neither is
    selected the status group imposes no restriction (this is what allows
    Overdue-only filtering)."""
    responded_on = config.get("responded") and is_customer_responded(t)
    waiting_on = config.get("waiting") and is_waiting_on_customer(t)
    if not (config.get("responded") or config.get("waiting")):
        return True  # no status restriction
    return responded_on or waiting_on


def matches_overdue(t, config):
    """Overdue is a separate AND condition combined with the selected status
    group. When Overdue is OFF it imposes no restriction."""
    return (not config.get("overdue")) or is_overdue(t)


def matches_missing_tags(t, config):
    """Missing Tags is a separate AND condition. When OFF it imposes no
    restriction."""
    return (not config.get("missing_tags")) or has_missing_tags(t)


def matches_days_window(t, config):
    """AND gate: ticket must have a valid updated_at within the last N days.
    Missing or malformed updated_at fails closed (excluded) so the window is
    never silently widened. Ticket count across fixtures is unaffected by the
    real clock because tests pin now_utc(); live data always carries
    updated_at from the list endpoint."""
    dt = parse_dt(t.get("updated_at"))
    if dt is None:
        return False
    cutoff = now_utc() - timedelta(days=int(config["days"]))
    return dt >= cutoff


def passes_filters(t, config=None):
    """Apply only the local filters the operator explicitly selected.

    There are no hidden queue gates here. In particular, status, subject
    keywords, Overdue, and Missing Tags do not restrict the cache unless their
    visible controls are selected. The two status controls OR together;
    Overdue and Missing Tags are independent AND restrictions when enabled.
    With every checkbox off this predicate returns True for every cached row.

    Days is not a local filter; it controls the Freshdesk retrieval window.
    Review View is applied separately from local SQLite state in ``queue()``.
    """
    cfg = config or dict(DEFAULT_FILTERS)
    if not matches_status_group(t, cfg):
        return False
    if not matches_overdue(t, cfg):
        return False
    if not matches_missing_tags(t, cfg):
        return False
    return True


# ---------------------------------------------------------------------------
# Freshdesk fetch (live only)
# ---------------------------------------------------------------------------


class QueueQuotaStop(Exception):
    """The current scan stopped before requesting another page."""


class QueueRateLimitError(Exception):
    """A bounded 429 retry sequence could not complete."""


class QueueCancelled(Exception):
    """Cooperative cancellation stopped the current queue retrieval."""


# Phase 2: Freshdesk's normal timestamp propagation is approximately one
# second in the audited data.  Five seconds is deliberately small; a longer
# unexplained tail remains UPDATED SINCE REVIEW.
CONVERSATION_UPDATE_TOLERANCE_SECONDS = 5
CONVERSATION_MAX_PAGES = 10


def conversation_classification(conversation):
    """Classify only defensible Freshdesk conversation metadata.

    ``incoming`` is the customer/external discriminator.  ``private`` is the
    internal-note discriminator; source and user_id are intentionally ignored.
    """
    if not isinstance(conversation, dict):
        return "ambiguous"
    incoming = conversation.get("incoming")
    private = conversation.get("private")
    if not isinstance(incoming, bool) or not isinstance(private, bool):
        return "ambiguous"
    if incoming and not private:
        return "customer"
    if not incoming and private:
        return "private"
    if not incoming and not private:
        return "public"
    return "ambiguous"


def _is_freshdesk_id(value):
    """Return true only for an unambiguous Freshdesk integer ID."""
    return isinstance(value, int) and not isinstance(value, bool)


def conversation_is_own_private_note(conversation, current_agent_id):
    """Whether an activity is proven to be the current agent's private note."""
    return (
        conversation_classification(conversation) == "private"
        and _is_freshdesk_id(current_agent_id)
        and _is_freshdesk_id(conversation.get("user_id"))
        and conversation["user_id"] == current_agent_id
    )


def fetch_current_agent_id(api_key):
    """Resolve the current Freshdesk agent once; ambiguity fails closed."""
    try:
        response = requests.get(
            f"https://{FRESHDESK_DOMAIN}/api/v2/agents/me",
            auth=(api_key, "X"), timeout=30,
        )
        if not 200 <= response.status_code < 300:
            return None
        payload = response.json()
    except Exception:
        return None
    if not isinstance(payload, dict):
        return None
    agent_id = payload.get("id")
    return agent_id if _is_freshdesk_id(agent_id) else None


def conversation_activity_timestamp(conversation):
    """Return max(valid created_at, updated_at), or None for malformed data."""
    if not isinstance(conversation, dict):
        return None
    values = []
    for key in ("created_at", "updated_at"):
        raw = conversation.get(key)
        if raw is not None:
            parsed = parse_dt(raw)
            if parsed is None:
                return None
            values.append(parsed)
    return max(values) if values else None


def review_ticket_fingerprint(ticket):
    """Canonicalize fields whose change is meaningful to review state.

    Tags are semantically unordered.  Volatile API bookkeeping fields are not
    included, so a private note can be the sole harmless change.
    """
    if not isinstance(ticket, dict):
        return None
    tags = ticket.get("tags")
    if not isinstance(tags, list):
        return None
    custom = ticket.get("custom_fields")
    if custom is not None and not isinstance(custom, dict):
        return None
    return (
        ticket.get("status"), ticket.get("subject"), ticket.get("priority"),
        ticket.get("type"), ticket.get("group_id"), ticket.get("responder_id"),
        ticket.get("due_by"), ticket.get("fr_due_by"),
        tuple(sorted(str(tag).strip().casefold() for tag in tags)),
        json.dumps(custom or {}, sort_keys=True, separators=(",", ":"), default=str),
    )


def _queue_settings():
    return (
        _safe_float_env("FRESHDESK_MIN_REQUEST_INTERVAL_SECONDS", 6, 1),
        _safe_int_env("FRESHDESK_MIN_REMAINING", 20, 0),
        min(_safe_int_env("FRESHDESK_MAX_RETRIES", 2, 0), 3),
    )


def _header_int(response, name):
    value = getattr(response, "headers", {}).get(name)
    try:
        parsed = int(value)
        return parsed if parsed >= 0 else None
    except (TypeError, ValueError):
        return None


def paginate_tickets(days=DAYS_DEFAULT, effective_since=None, clock=None, sleeper=None,
                      progress_callback=None, cancel_callback=None):
    """Fetch all tickets conservatively, optionally reporting progress.

    Request count means Freshdesk HTTP request attempts, including 429s.
    Existing pacing, retry, quota, and GET-only behavior are preserved.
    """
    days = parse_days(days)
    clock = clock or time.monotonic
    sleeper = sleeper or time.sleep
    started = clock()
    interval, min_remaining, max_retries = _queue_settings()
    page = 1
    per_page = 100
    last_start = None
    pages_completed = 0
    tickets_received = 0
    request_count = 0
    last_remaining = None
    def emit(state="running", wait_seconds=0, error=None):
        if progress_callback:
            progress_callback({"state": state, "page": page,
                               "pages_completed": pages_completed,
                               "tickets_received": tickets_received,
                               "request_count": request_count,
                               "rate_limit_remaining": last_remaining,
                               "elapsed_seconds": round(clock() - started, 3),
                               "wait_seconds": wait_seconds, "error": error})
    def wait(seconds):
        remaining = max(0, seconds)
        if remaining:
            emit("waiting", remaining)
        if not cancel_callback:
            sleeper(remaining)
            return True
        while remaining > 0:
            if cancel_callback():
                emit("cancelled")
                return False
            step = min(remaining, 0.25)
            sleeper(step)
            remaining -= step
        return not cancel_callback()
    since = effective_since or (now_utc() - timedelta(days=days)).isoformat()
    while True:
        if cancel_callback and cancel_callback():
            emit("cancelled")
            raise QueueCancelled("Queue refresh cancelled.")
        retries = 0
        while True:
            elapsed = None if last_start is None else clock() - last_start
            if elapsed is not None and elapsed < interval and not wait(interval - elapsed):
                raise QueueCancelled("Queue refresh cancelled.")
            last_start = clock()
            emit("before_request")
            url = f"https://{FRESHDESK_DOMAIN}/api/v2/tickets"
            params = {"page": page, "per_page": per_page, "updated_since": since}
            request_count += 1
            r = requests.get(url, auth=fd_auth(), params=params, timeout=30)
            if r.status_code != 429:
                break
            retry_after = _header_int(r, "Retry-After")
            wait_seconds = retry_after if retry_after is not None else max(6, interval)
            if retries >= max_retries:
                emit("failed", error=f"Freshdesk rate limit retry limit reached on page {page}.")
                raise QueueRateLimitError(
                    f"Freshdesk rate limit retry limit reached on page {page}."
                )
            if not wait(wait_seconds):
                raise QueueCancelled("Queue refresh cancelled.")
            retries += 1

        r.raise_for_status()
        data = r.json()
        # Read supported quota headers after every response. Invalid or missing
        # metadata is intentionally ignored; only a valid Remaining value can
        # activate the conservative safety stop.
        rate_limit_metadata = {
            name: _header_int(r, name)
            for name in (
                "X-RateLimit-Total",
                "X-RateLimit-Remaining",
                "X-RateLimit-Used-CurrentRequest",
                "Retry-After",
            )
        }
        remaining = rate_limit_metadata["X-RateLimit-Remaining"]
        last_remaining = remaining
        emit("response")
        if not data:
            break
        for ticket in data:
            if cancel_callback and cancel_callback():
                emit("cancelled")
                raise QueueCancelled("Queue refresh cancelled.")
            tickets_received += 1
            yield ticket
        pages_completed += 1
        emit("running")
        if len(data) < per_page:
            break
        if remaining is not None and remaining <= min_remaining:
            raise QueueQuotaStop(
                f"Freshdesk API quota is getting low. Retrieval stopped with {remaining} calls remaining."
            )
        page += 1


def offline_paginate_tickets():
    """Yield tickets from local fixture pages. Offline mode only.

    Fail closed: a missing file, malformed JSON, or a fixture with the wrong
    shape raises OfflineDataError. This function never touches the network and
    never reads the API key file.
    """
    try:
        with open(FIXTURES_FILE, "r") as fh:
            data = json.load(fh)
    except FileNotFoundError:
        raise OfflineDataError(
            f"Offline mode: fixture file not found at {FIXTURES_FILE}. "
            "Run the scanner with live mode, or restore fixtures/fixtures.json."
        )
    except json.JSONDecodeError as e:
        raise OfflineDataError(
            f"Offline mode: fixture file {FIXTURES_FILE} is malformed JSON ({e})."
        )
    if not isinstance(data, dict) or not isinstance(data.get("pages"), list) or not data["pages"]:
        raise OfflineDataError(
            'Offline mode: fixture file must contain a non-empty "pages" list of ticket lists.'
        )
    for page in data["pages"]:
        if not isinstance(page, list):
            raise OfflineDataError(
                "Offline mode: each fixture page must be a list of tickets."
            )
        for t in page:
            if not isinstance(t, dict):
                raise OfflineDataError(
                    "Offline mode: each fixture ticket must be a JSON object."
                )
            yield t


def queue_cache_timestamp(value=None):
    """Return a canonical cache-metadata timestamp: UTC, whole seconds, trailing Z."""
    value = now_utc() if value is None else value
    if not isinstance(value, datetime) or value.tzinfo is None or value.utcoffset() is None:
        raise ValueError("Queue cache timestamps must be timezone-aware datetimes.")
    return value.astimezone(timezone.utc).replace(microsecond=0).strftime("%Y-%m-%dT%H:%M:%SZ")


def _valid_queue_cache_timestamp(value):
    if not isinstance(value, str):
        return False
    try:
        parsed = datetime.strptime(value, "%Y-%m-%dT%H:%M:%SZ")
    except ValueError:
        return False
    return queue_cache_timestamp(parsed.replace(tzinfo=timezone.utc)) == value


def _valid_queue_cache_days(value):
    return isinstance(value, int) and not isinstance(value, bool) and DAYS_MIN <= value <= DAYS_MAX


def normalize_live_queue_cache(blob):
    """Validate one queue-cache envelope without mutating or migrating it.

    A schema-less envelope is legacy: its metadata remains explicitly unknown.
    Unsupported versions and malformed v2 envelopes fail closed (``None``).
    """
    if not isinstance(blob, dict) or not isinstance(blob.get("tickets"), list):
        return None
    schema_version = blob.get("schema_version")
    if schema_version is None:
        return {**blob, "cache_metadata": {
            "schema_version": None,
            "last_successful_refresh_started_at": None,
            "last_successful_refresh_finished_at": None,
            "last_refresh_mode": "legacy",
            "last_refresh_requested_days": None,
            "rolling_retention_days": None,
        }}
    if (not isinstance(schema_version, int) or isinstance(schema_version, bool)
            or schema_version != QUEUE_CACHE_SCHEMA_VERSION):
        return None
    # A v2 envelope without an established cursor is structurally safe to
    # reconcile as a baseline; only the cursor itself is unknown.
    required = (
        "days", "fetched_at", "last_refresh_mode",
        "last_refresh_requested_days", "rolling_retention_days",
    )
    if any(key not in blob for key in required):
        return None
    started = blob.get("last_successful_refresh_started_at")
    finished = blob.get("last_successful_refresh_finished_at")
    if (not _valid_queue_cache_days(blob["days"])
            or not isinstance(blob["fetched_at"], (int, float))
            or isinstance(blob["fetched_at"], bool)
            or (started is not None and not isinstance(started, str))
            or (finished is not None and not isinstance(finished, str))
            or blob["last_refresh_mode"] not in QUEUE_CACHE_REFRESH_MODES
            or blob["last_refresh_mode"] == "legacy"
            or not _valid_queue_cache_days(blob["last_refresh_requested_days"])
            or not isinstance(blob["rolling_retention_days"], int)
            or isinstance(blob["rolling_retention_days"], bool)
            or blob["rolling_retention_days"] <= 0):
        return None
    return {**blob, "cache_metadata": {
        "schema_version": blob["schema_version"],
        "last_successful_refresh_started_at": started,
        "last_successful_refresh_finished_at": finished,
        "last_refresh_mode": blob["last_refresh_mode"],
        "last_refresh_requested_days": blob["last_refresh_requested_days"],
        "rolling_retention_days": blob["rolling_retention_days"],
    }}


def queue_refresh_plan(cache_blob, days, attempt_started_at, mode=None):
    """Choose the retrieval horizon for one normal or explicit reconcile attempt.

    Normal refreshes use the persistent cursor when safe, otherwise the fixed
    60-day initialization baseline. Reconcile deliberately bypasses the cursor
    for this attempt only; both paths still merge, analyze, retain, and save
    atomically through the same worker.
    """
    requested_days = parse_days(days)
    legacy_plan = mode is None
    mode = "normal" if legacy_plan else mode
    if mode not in {"normal", "reconcile"}:
        raise ValueError("Unknown queue refresh mode.")
    days = requested_days if mode == "reconcile" or legacy_plan else DAYS_DEFAULT
    attempt_started_at = attempt_started_at.astimezone(timezone.utc).replace(microsecond=0)
    # Schema-v2 cache metadata is durable at the top level.  The loader
    # synthesizes cache_metadata only for legacy/compatibility callers, so it
    # may fill an absent top-level field but must never override one.
    envelope = cache_blob if isinstance(cache_blob, dict) else {}
    compatibility_metadata = envelope.get("cache_metadata")
    if not isinstance(compatibility_metadata, dict):
        compatibility_metadata = {}

    def planning_metadata(field):
        return (envelope[field] if field in envelope
                else compatibility_metadata.get(field))

    schema_version = planning_metadata("schema_version")
    raw_cursor = planning_metadata("last_successful_refresh_started_at")
    raw_finished = planning_metadata("last_successful_refresh_finished_at")
    cursor = parse_dt(raw_cursor)
    finished = parse_dt(raw_finished)
    # A cursor later than this attempt is future metadata, including after a
    # local clock rollback. Never derive an incremental horizon from it: the
    # normal Days baseline reconciles conservatively and replaces it only after
    # successful atomic cache commit.
    cursor_not_future = cursor is not None and cursor <= attempt_started_at
    valid_cursor = (
        schema_version == QUEUE_CACHE_SCHEMA_VERSION
        and _valid_queue_cache_timestamp(raw_cursor)
        and _valid_queue_cache_timestamp(raw_finished)
        and cursor is not None and finished is not None
        and cursor <= finished and cursor_not_future
    )
    # Corrupt timestamp strings are an invalid cursor, not an unsafe ticket
    # envelope: reconcile safely from the requested Days baseline.  Non-string
    # timestamp metadata is rejected earlier by envelope validation.
    if (raw_cursor is not None and not isinstance(raw_cursor, str)) or (
            raw_finished is not None and not isinstance(raw_finished, str)):
        raise ValueError("Queue cursor metadata is structurally unsafe.")
    if mode == "reconcile":
        return {
            "refresh_mode": "reconcile", "cursor_source": "requested_days",
            "effective_updated_since": queue_cache_timestamp(attempt_started_at - timedelta(days=days)),
            "durable_refresh_started_at": attempt_started_at,
        }
    if valid_cursor:
        cursor = cursor.astimezone(timezone.utc).replace(microsecond=0)
        effective = cursor - timedelta(seconds=INCREMENTAL_CURSOR_OVERLAP_SECONDS)
        return {
            "refresh_mode": "incremental", "cursor_source": "previous_successful_start",
            "effective_updated_since": queue_cache_timestamp(effective),
            "durable_refresh_started_at": max(cursor, attempt_started_at),
        }
    return {
        "refresh_mode": "baseline", "cursor_source": "days_baseline",
        "effective_updated_since": queue_cache_timestamp(attempt_started_at - timedelta(days=days)),
        "durable_refresh_started_at": attempt_started_at,
    }


def load_live_queue_cache():
    """Read and validate the single LIVE queue-cache envelope without rewriting it."""
    if not os.path.exists(LIVE_QUEUE_CACHE_FILE):
        return None
    try:
        with open(LIVE_QUEUE_CACHE_FILE, "r") as fh:
            blob = json.load(fh)
    except (OSError, ValueError):
        return None
    return normalize_live_queue_cache(blob)


def save_live_queue_cache(tickets, days=DAYS_DEFAULT, refresh_started_at=None, refresh_finished_at=None,
                           refresh_mode="reconcile", effective_updated_since=None):
    """Atomically save a successfully reconciled queue-cache envelope.

    ``fetched_at`` remains a compatibility Unix timestamp and denotes the same
    completion instant as ``last_successful_refresh_finished_at``.
    """
    days = parse_days(days)
    started = queue_cache_timestamp(refresh_started_at)
    finished_dt = now_utc() if refresh_finished_at is None else refresh_finished_at
    finished = queue_cache_timestamp(finished_dt)
    payload = {
        "schema_version": QUEUE_CACHE_SCHEMA_VERSION,
        "days": days,
        "fetched_at": finished_dt.timestamp(),
        "last_successful_refresh_started_at": started,
        "last_successful_refresh_finished_at": finished,
        "last_refresh_mode": refresh_mode,
        "last_refresh_requested_days": days,
        "rolling_retention_days": ROLLING_RETENTION_DAYS,
        "tickets": list(tickets),
    }
    if effective_updated_since is not None:
        payload["last_refresh_effective_updated_since"] = effective_updated_since
    directory = os.path.dirname(os.path.abspath(LIVE_QUEUE_CACHE_FILE)) or "."
    os.makedirs(directory, exist_ok=True)
    fd, tmp = tempfile.mkstemp(prefix=".queue_live_tickets.", suffix=".tmp", dir=directory)
    try:
        with os.fdopen(fd, "w") as fh:
            json.dump(payload, fh, indent=2, sort_keys=True)
            fh.flush()
            os.fsync(fh.fileno())
        os.replace(tmp, LIVE_QUEUE_CACHE_FILE)
    except Exception:
        try:
            os.unlink(tmp)
        except OSError:
            pass
        raise


def fetch_live_queue(days=DAYS_DEFAULT, effective_since=None, api_key=None, progress_callback=None,
                     cancel_callback=None):
    """Perform exactly one manual Freshdesk queue retrieval from an explicit horizon."""
    return list(paginate_tickets(days=parse_days(days), effective_since=effective_since,
                                 progress_callback=progress_callback,
                                 cancel_callback=cancel_callback))


def fetch_ticket_conversations(ticket_id, reviewed_at, api_key=None, clock=None,
                                sleeper=None, cancel_callback=None,
                                progress_callback=None):
    """Fetch just enough conversation pages to cross ``reviewed_at``.

    This uses the same interval/quota/retry policy as ticket retrieval.  A
    bounded page ceiling or any malformed/failed response is inconclusive.
    Returns ``(conversations, complete, remaining)``.
    """
    reviewed = parse_dt(reviewed_at)
    if reviewed is None:
        return [], False, None
    clock = clock or time.monotonic
    sleeper = sleeper or time.sleep
    interval, min_remaining, max_retries = _queue_settings()
    last_start = None
    conversations = []
    remaining = None
    for page in range(1, CONVERSATION_MAX_PAGES + 1):
        if cancel_callback and cancel_callback():
            return conversations, False, remaining
        elapsed = None if last_start is None else clock() - last_start
        if elapsed is not None and elapsed < interval:
            wait = interval - elapsed
            if cancel_callback:
                end = clock() + wait
                while clock() < end:
                    if cancel_callback():
                        return conversations, False, remaining
                    sleeper(min(0.25, end - clock()))
            else:
                sleeper(wait)
        last_start = clock()
        retries = 0
        while True:
            if cancel_callback and cancel_callback():
                return conversations, False, remaining
            url = f"https://{FRESHDESK_DOMAIN}/api/v2/tickets/{int(ticket_id)}/conversations"
            try:
                response = requests.get(url, auth=fd_auth(),
                                        params={"page": page, "per_page": 30}, timeout=30)
            except Exception:
                return conversations, False, remaining
            if response.status_code != 429:
                break
            retry_after = _header_int(response, "Retry-After")
            wait = retry_after if retry_after is not None else max(6, interval)
            if retries >= max_retries:
                return conversations, False, remaining
            if cancel_callback:
                end = clock() + wait
                while clock() < end:
                    if cancel_callback():
                        return conversations, False, remaining
                    sleeper(min(0.25, end - clock()))
            else:
                sleeper(wait)
            retries += 1
        try:
            response.raise_for_status()
            data = response.json()
        except (requests.RequestException, ValueError):
            return conversations, False, remaining
        remaining = _header_int(response, "X-RateLimit-Remaining")
        if remaining is not None and remaining <= min_remaining:
            # This response is usable, but no further request is safe.
            if not isinstance(data, list):
                return conversations, False, remaining
            conversations.extend(data)
            oldest = [conversation_activity_timestamp(c) for c in data]
            if any(value is None for value in oldest) or not any(value <= reviewed for value in oldest):
                return conversations, False, remaining
            return conversations, True, remaining
        if not isinstance(data, list):
            return conversations, False, remaining
        conversations.extend(data)
        timestamps = [conversation_activity_timestamp(c) for c in data]
        if any(value is None for value in timestamps):
            return conversations, False, remaining
        if any(value <= reviewed for value in timestamps):
            return conversations, True, remaining
        if len(data) < 30:
            return conversations, True, remaining
    return conversations, False, remaining


def _backup_after_mutation(reason):
    try:
        review_backups.create_backup(reason=reason)
        return None
    except Exception as exc:
        review_backups.LOGGER.error("Local review-state backup failed after %s: %s", reason, exc, exc_info=True)
        return "Review saved, but local backup failed."


_STARTUP_BACKUP_DONE = False


@app.before_request
def _protect_review_state_at_startup():
    global _STARTUP_BACKUP_DONE
    if _STARTUP_BACKUP_DONE:
        return
    try:
        review_backups.ensure_startup_backup()
    except Exception:
        review_backups.LOGGER.error("Startup review-state backup failed", exc_info=True)
    _STARTUP_BACKUP_DONE = True


def _advance_review_snapshot(ticket_id, updated_at, backup=True):
    """Advance one existing queue review snapshot, preserving every field."""
    if parse_dt(updated_at) is None:
        return False
    conn = _db_conn()
    try:
        row = conn.execute(
            "SELECT review_result FROM review_state WHERE ticket_id = ?", (ticket_id,)
        ).fetchone()
        if row is None or row["review_result"] not in REVIEWED_STATES:
            return False
        conn.execute(
            "UPDATE review_state SET reviewed_updated_at = ?, modified_at = ? WHERE ticket_id = ?",
            (updated_at, iso_now(), ticket_id),
        )
        conn.commit()
        if backup:
            _backup_after_mutation("automatic-review-advance")
        return True
    finally:
        conn.close()


def _reconcile_queue_refresh(old_blob, incoming_tickets, progress_callback=None,
                               cancel_callback=None, attempt_started_at=None,
                               current_agent_id_fetcher=None):
    """Merge, analyze effective changes, retain, then defer review snapshots.

    Retention reads local review state only.  Its retained ticket list is the
    sole cache-save input; review rows are neither created nor removed here.
    """
    existing_tickets = old_blob["tickets"] if old_blob else []
    merged = merge_queue_tickets(existing_tickets, incoming_tickets)
    old_by_id = {ticket["id"]: ticket for ticket in existing_tickets}
    effective_changes = [
        ticket for ticket in merged.tickets
        if old_by_id.get(ticket["id"]) != ticket
    ]
    _, finalize_after_save = _prepare_conversation_review_updates(
        effective_changes, old_blob, progress_callback=progress_callback,
        cancel_callback=cancel_callback,
        current_agent_id_fetcher=current_agent_id_fetcher,
    )
    # Load once, read-only, after conversation preparation. The pure engine
    # validates the complete merged cache and canonical state mapping before it
    # can classify any object. Attempt start is supplied by the refresh job.
    review_states = {
        ticket_id: row.get("review_result")
        for ticket_id, row in load_review_rows().items()
    }
    retention = apply_queue_retention(
        merged.tickets, review_states,
        reference_time=attempt_started_at if attempt_started_at is not None else now_utc(),
    )
    retained_ids = {ticket["id"] for ticket in retention.tickets}

    def finalize_retained_after_save():
        # A snapshot advance for an object omitted from the durable cache would
        # falsely imply that its corresponding cached state was preserved.
        # Keep it conservatively stale instead; its review row remains intact.
        if finalize_after_save:
            finalize_after_save(retained_ids)

    return retention.tickets, finalize_retained_after_save, merged.metrics, retention.metrics



def _prepare_conversation_review_updates(tickets, old_blob, progress_callback=None,
                                           cancel_callback=None,
                                           current_agent_id_fetcher=None):
    """Classify narrow reviewed/newer candidates before cache commit."""
    old_tickets = (old_blob or {}).get("tickets") if isinstance(old_blob, dict) else None
    old_by_id = {t.get("id"): t for t in old_tickets or [] if isinstance(t, dict)}
    states = load_review_rows()
    candidates = []
    for ticket in tickets:
        tid = ticket.get("id") if isinstance(ticket, dict) else None
        state = states.get(tid)
        reviewed = state.get("reviewed_updated_at") if state else None
        if state and state.get("review_result") in REVIEWED_STATES:
            if parse_dt(reviewed) and parse_dt(ticket.get("updated_at")) and parse_dt(ticket.get("updated_at")) > parse_dt(reviewed):
                candidates.append(ticket)
    updates = {}
    inconclusive = 0
    identity_resolved = False
    current_agent_id = None
    for index, ticket in enumerate(candidates, 1):
        if cancel_callback and cancel_callback():
            break
        tid = ticket["id"]
        state = states.get(tid)
        old = old_by_id.get(tid)
        if (old is None or review_ticket_fingerprint(old) is None or
                review_ticket_fingerprint(ticket) != review_ticket_fingerprint(old)):
            inconclusive += 1
            continue
        if progress_callback:
            progress_callback({"current_stage": "Checking reviewed updates",
                               "conversation_candidates": len(candidates),
                               "conversation_checks_completed": index,
                               "private_note_updates_suppressed": len(updates),
                               "conversation_checks_inconclusive": inconclusive})
        conversations, complete, remaining = fetch_ticket_conversations(
            tid, state["reviewed_updated_at"], cancel_callback=cancel_callback,
            progress_callback=progress_callback)
        if not complete:
            inconclusive += 1
            continue
        review_at = parse_dt(state["reviewed_updated_at"])
        ticket_at = parse_dt(ticket.get("updated_at"))
        post = [
            (conversation, activity_at)
            for conversation in conversations
            for activity_at in [conversation_activity_timestamp(conversation)]
            if activity_at is not None and activity_at > review_at
        ]
        if not post:
            inconclusive += 1
            continue
        if not identity_resolved:
            identity_resolved = True
            try:
                current_agent_id = current_agent_id_fetcher() if current_agent_id_fetcher else None
            except Exception:
                current_agent_id = None
        if any(not conversation_is_own_private_note(c, current_agent_id) for c, _ in post):
            inconclusive += 1
            continue
        latest = max(value for _, value in post)
        if ticket_at < latest or (ticket_at - latest).total_seconds() > CONVERSATION_UPDATE_TOLERANCE_SECONDS:
            inconclusive += 1
            continue
        updates[tid] = ticket.get("updated_at")
    if progress_callback:
        progress_callback({"current_stage": "Checking reviewed updates",
                           "conversation_candidates": len(candidates),
                           "conversation_checks_completed": len(candidates),
                           "private_note_updates_suppressed": len(updates),
                           "conversation_checks_inconclusive": inconclusive})
    def apply_updates(retained_ids=None):
        changed = False
        for tid, timestamp in updates.items():
            if retained_ids is not None and tid not in retained_ids:
                continue
            try:
                changed = _advance_review_snapshot(tid, timestamp, backup=False) or changed
            except Exception:
                # Cache remains valid; this ticket simply remains flagged.
                continue
        if changed:
            _backup_after_mutation("automatic-review-advance")
    return list(tickets), apply_updates


def get_ticket_pool():
    """Return (raw_tickets, cache_age_seconds) — RENDER ONLY, never fetches.

    Offline mode loads fixtures; live mode reads only the live cache. Missing,
    stale, or legacy cache state never triggers a fetch.
    """
    if is_offline() and os.environ.get("FRESHDESK_OFFLINE_CACHE", "").strip().lower() not in ("1", "true", "yes"):
        raw = list(offline_paginate_tickets())
        return raw, None

    # Live mode and explicit offline-cache preview mode: cache read only. No
    # network, no auto-refresh, no fetch.

    blob = load_live_queue_cache()
    if blob:
        raw = blob["tickets"]
        cache_age = now_utc().timestamp() - blob.get("fetched_at", now_utc().timestamp())
        cached_days = blob.get("days") if isinstance(blob.get("days"), int) else None
        return raw, int(cache_age)
    # No live cache yet: render an empty pool. GET must not fetch.
    return [], None


def apply_queue_filters(tickets, config):
    """Apply default Review Scope first, then opt-in local queue filters, and
    dedupe by ticket id.

    The cache already represents the selected Freshdesk retrieval window, so
    Days is deliberately *not* applied again here. Review Scope runs first
    (photo/video subjects only + reviewed/closed tag exclusions); the manual
    filters may narrow that result further. With the Review Scope controls and
    every manual control off, this returns the complete cached ticket list
    (deduped only).
    """
    seen = set()
    out = []
    for t in tickets:
        tid = t.get("id")
        if tid in seen:
            continue
        seen.add(tid)
        if not passes_review_scope(t, config):
            continue
        if passes_filters(t, config):
            out.append(t)
    return out


# ---------------------------------------------------------------------------
# Local review state (SQLite, never sent to Freshdesk)
# ---------------------------------------------------------------------------

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS review_state (
    ticket_id           INTEGER PRIMARY KEY,
    review_result       TEXT    NOT NULL DEFAULT 'Unreviewed',
    first_opened_at     TEXT,
    last_opened_at      TEXT,
    last_review_change_at TEXT,
    reviewed_updated_at TEXT,
    note                TEXT    NOT NULL DEFAULT '',
    created_at          TEXT    NOT NULL,
    modified_at         TEXT    NOT NULL
);
"""

# Separate namespace for closed-ticket housekeeping reviews (Prompt 12). The
# /closed page must never read or write /queue review_state rows. Columns mirror
# review_state plus a closed_at snapshot of the source ticket.
CLOSED_SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS closed_review_state (
    ticket_id           INTEGER PRIMARY KEY,
    review_result       TEXT    NOT NULL DEFAULT 'Unreviewed',
    first_opened_at     TEXT,
    last_opened_at      TEXT,
    last_review_change_at TEXT,
    reviewed_updated_at TEXT,
    closed_at_snapshot  TEXT,
    note                TEXT    NOT NULL DEFAULT '',
    created_at          TEXT    NOT NULL,
    modified_at         TEXT    NOT NULL
);
"""

# Tables the local-review SQL may target. The queue table and the closed table
# stay fully separate; only the SQL layer is shared (parameterized statements
# are always value-bound, and the table name comes from this fixed allowlist).
_REVIEW_TABLES = ("review_state", "closed_review_state")


def get_db_path():
    return os.environ.get("REVIEW_DB_PATH") or os.path.join(BASE_DIR, "data", "review_state.sqlite3")


def init_db(path=None):
    """Create the review-state databases (and parent dir) if missing. Both
    namespaces (queue review_state and closed_review_state) live in the same
    SQLite file; each table is created idempotently. Used by tests and
    validate.sh with a temporary path; callers never need a live database for
    offline development."""
    db_path = path or get_db_path()
    parent = os.path.dirname(db_path)
    if parent:
        os.makedirs(parent, exist_ok=True)
    conn = sqlite3.connect(db_path)
    try:
        conn.execute(SCHEMA_SQL)
        conn.execute(CLOSED_SCHEMA_SQL)
        conn.commit()
    finally:
        conn.close()
    return db_path


def _db_conn():
    db_path = get_db_path()
    parent = os.path.dirname(db_path)
    if parent:
        os.makedirs(parent, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute(SCHEMA_SQL)
    conn.execute(CLOSED_SCHEMA_SQL)
    conn.commit()
    return conn


def load_review_rows():
    """Return {ticket_id: row-dict} for every stored queue review state."""
    conn = _db_conn()
    try:
        rows = conn.execute("SELECT * FROM review_state").fetchall()
        return {r["ticket_id"]: dict(r) for r in rows}
    finally:
        conn.close()


def load_closed_review_rows():
    """Return {ticket_id: row-dict} for every stored closed-ticket review
    state. This is the /closed page's own namespace — queue review_state rows
    are never read here."""
    conn = _db_conn()
    try:
        rows = conn.execute("SELECT * FROM closed_review_state").fetchall()
        return {r["ticket_id"]: dict(r) for r in rows}
    finally:
        conn.close()


def _last_opened_base(table: str):
    """Shared Last Opened selection over one table (queue or closed).

    The Last Opened focus marker is derived from the newest valid
    `last_opened_at` across the table's rows (spec §3/§6) — never from the
    review result, so a ticket can be e.g. \"Resolved + Last Opened\" at the
    same time. Selection is deterministic:

    - every stored last_opened_at is parsed as UTC; missing, empty, or
      malformed values are skipped (fail safe);
    - the chosen ticket is the maximum (last_opened_at, ticket_id), so two
      records with identical timestamps resolve to the higher ticket id
      instead of leaving the marker ambiguous;
    - None is returned when no ticket has ever been opened, in which case
      the dashboard renders neither marker nor jump control.
    """
    conn = _db_conn()
    try:
        rows = conn.execute(
            f"SELECT ticket_id, last_opened_at FROM {table}"
        ).fetchall()
    finally:
        conn.close()
    best = None  # (parsed UTC datetime, ticket_id)
    for r in rows:
        ts = parse_dt(r["last_opened_at"])
        if ts is None:
            continue  # missing / malformed timestamp fails safe
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)  # naive stored value == UTC
        key = (ts, r["ticket_id"])
        if best is None or key > best:
            best = key
    return best[1] if best else None


def last_opened_ticket_id():
    """Ticket id of the single most recently opened QUEUE ticket, or None."""
    return _last_opened_base("review_state")


def closed_last_opened_ticket_id():
    """Ticket id of the single most recently opened CLOSED ticket, or None."""
    return _last_opened_base("closed_review_state")


def _state_row(table: str, ticket_id):
    conn = _db_conn()
    try:
        return conn.execute(
            f"SELECT * FROM {table} WHERE ticket_id = ?", (ticket_id,)
        ).fetchone()
    finally:
        conn.close()


def _mark_opened(table: str, ticket_id):
    """Record that a ticket link was opened. Pure local state — no Freshdesk
    interaction. Shared by the /queue and /closed pages; each page passes its
    own table so the namespaces never mix.

    Review-result rule (dashboard spec §8): an Unreviewed ticket becomes
    \"Opened / In Review\"; an already-opened ticket stays opened; deliberate
    states (Resolved, Not Applicable to Me, No Action Needed, Needs Follow-Up)
    are PRESERVED — re-opening a link must never silently erase a deliberate
    review result. first_opened_at is set once, last_opened_at always updates,
    and no duplicate record is ever created.

    Returns the effective review_result for the ticket after marking.
    """
    now = iso_now()
    conn = _db_conn()
    try:
        row = conn.execute(
            f"SELECT * FROM {table} WHERE ticket_id = ?", (ticket_id,)
        ).fetchone()
        if row is None:
            conn.execute(
                f"INSERT INTO {table} (ticket_id, review_result, first_opened_at, last_opened_at,"
                " last_review_change_at, reviewed_updated_at, note, created_at, modified_at)"
                " VALUES (?,?,?,?,?,?,?,?,?)",
                (ticket_id, "Opened / In Review", now, now, now, None, "", now, now),
            )
            result = "Opened / In Review"
        else:
            current = row["review_result"]
            first = row["first_opened_at"] or now
            if not current or current == "Unreviewed":
                result = "Opened / In Review"
            else:
                result = current  # preserve deliberate/completed states
            changed = result != current
            conn.execute(
                f"UPDATE {table} SET review_result = ?, first_opened_at = ?,"
                " last_opened_at = ?, last_review_change_at = ?, modified_at = ? WHERE ticket_id = ?",
                (result, first, now, now if changed else row["last_review_change_at"], now, ticket_id),
            )
        conn.commit()
        _backup_after_mutation("closed-review-change" if table == "closed_review_state" else "review-change")
        return result
    finally:
        conn.close()


def mark_opened(ticket_id):
    """Record that a QUEUE ticket link was opened (queue namespace)."""
    return _mark_opened("review_state", ticket_id)


def mark_closed_opened(ticket_id):
    """Record that a CLOSED ticket link was opened (closed namespace)."""
    return _mark_opened("closed_review_state", ticket_id)


def _set_review_result(table: str, ticket_id, result, reviewed_updated_at=None):
    """Shared local review-result save. `reviewed_updated_at` snapshots the
    ticket's updated_at at review time for the Reviewed states; for Unreviewed
    / Opened the snapshot is cleared so no stale flag can linger."""
    if table not in _REVIEW_TABLES:
        raise ValueError(f"unknown review table: {table!r}")
    if result not in REVIEW_STATES:
        raise ValueError(f"unknown review result: {result!r}")
    now = iso_now()
    conn = _db_conn()
    try:
        row = conn.execute(f"SELECT * FROM {table} WHERE ticket_id = ?", (ticket_id,)).fetchone()
        if row is None:
            conn.execute(
                f"INSERT INTO {table} (ticket_id, review_result, first_opened_at, last_opened_at,"
                " last_review_change_at, reviewed_updated_at, note, created_at, modified_at)"
                " VALUES (?,?,?,?,?,?,?,?,?)",
                (ticket_id, result, None, None, now,
                 reviewed_updated_at if result in REVIEWED_STATES else None, "", now, now),
            )
        else:
            conn.execute(
                f"UPDATE {table} SET review_result = ?, last_review_change_at = ?,"
                " reviewed_updated_at = ?, modified_at = ? WHERE ticket_id = ?",
                (result, now,
                 reviewed_updated_at if result in REVIEWED_STATES else None,
                 now, ticket_id),
            )
        conn.commit()
        _backup_after_mutation("closed-review-change" if table == "closed_review_state" else "review-change")
    finally:
        conn.close()


def set_review_result(ticket_id, result, reviewed_updated_at=None):
    """Save a local QUEUE review result (queue namespace)."""
    _set_review_result("review_state", ticket_id, result, reviewed_updated_at)


def set_closed_review_result(ticket_id, result, reviewed_updated_at=None):
    """Save a local CLOSED review result (closed namespace)."""
    _set_review_result("closed_review_state", ticket_id, result, reviewed_updated_at)


# ---------------------------------------------------------------------------
# Review view logic (Active / Completed / All)
# ---------------------------------------------------------------------------


def updated_since_review(t, state_row):
    """True when the ticket's current updated_at is strictly newer than the
    updated_at snapshot taken when a Reviewed state was assigned. Fail-safe:
    malformed or missing timestamps never produce the flag."""
    if state_row is None:
        return False
    snapshot = state_row.get("reviewed_updated_at")
    if not snapshot:
        return False
    ticket_dt = parse_dt(t.get("updated_at"))
    snapshot_dt = parse_dt(snapshot)
    if ticket_dt is None or snapshot_dt is None:
        return False
    return ticket_dt > snapshot_dt


def review_view_includes(state_row, updated_flag, view):
    """Which tickets belong to the requested review view.

    active:    Unreviewed, Opened / In Review, Needs Follow-Up, or any ticket
               flagged UPDATED SINCE REVIEW (even a previously Completed one).
    completed: Resolved, Not Applicable to Me, No Action Needed — but only if
               NOT flagged, because an updated ticket returns to Active.
    all:       everything.
    """
    if view == "all":
        return True
    result = state_row.get("review_result", "Unreviewed") if state_row else "Unreviewed"
    if updated_flag:
        return view == "active"
    if view == "active":
        return result in ACTIVE_STATES
    return result in COMPLETED_STATES


def closed_review_view_includes(state_row, view):
    """Closed-housekeeping review view membership (local only).

    Active:    Unreviewed, Opened / In Review, Needs Follow-Up.
    Completed: Resolved, Not Applicable to Me, No Action Needed.
    All:       every local closed review state.

    Always reflects the CLOSED review namespace (Prompt 12). The queue's
    "UPDATED SINCE REVIEW" bounce-back does not apply here: /closed shows
    housekeeping review state only, without the queue's updated_at flag.
    """
    if view == "all":
        return True
    result = state_row.get("review_result", "Unreviewed") if state_row else "Unreviewed"
    if view == "active":
        return result in ACTIVE_STATES
    return result in COMPLETED_STATES


# ---------------------------------------------------------------------------
# Badges
# ---------------------------------------------------------------------------

REVIEW_CLASS = {
    "Unreviewed": "rv-unreviewed",
    "Opened / In Review": "rv-opened",
    "Resolved": "rv-resolved",
    "Not Applicable to Me": "rv-na",
    "No Action Needed": "rv-none",
    "Needs Follow-Up": "rv-followup",
}


def sla_unavailable(t):
    """Customer-responded ticket whose due_by is missing or malformed: the SLA
    date simply is not available, so the dashboard says so instead of guessing."""
    if t.get("status") != 2:
        return False
    due = t.get("due_by")
    if not due:
        return True
    return parse_dt(due) is None


def ticket_badges(t, state_row, updated_flag):
    """All badges for one ticket row. Text + CSS class pair; never color alone
    (screen readers and color-blind users get the text)."""
    badges = []
    result = state_row.get("review_result", "Unreviewed") if state_row else "Unreviewed"
    # Display text: the Opened state renders as the all-caps "OPENED / IN REVIEW"
    # badge (spec §3.5/§7) while the select option value stays sentence case.
    badge_text = "OPENED / IN REVIEW" if result == "Opened / In Review" else result
    badges.append(("review", badge_text, "b-review " + REVIEW_CLASS.get(result, "rv-unreviewed")))
    if is_overdue(t):
        badges.append(("attr", "OVERDUE", "b-overdue"))
    if is_customer_responded(t):
        badges.append(("attr", "CUSTOMER RESPONDED", "b-responded"))
    if is_waiting_on_customer(t):
        badges.append(("attr", "WAITING ON CUSTOMER", "b-waiting"))
    if has_missing_tags(t):
        badges.append(("attr", "MISSING TAGS", "b-missing"))
    if sla_unavailable(t):
        badges.append(("attr", "SLA DATE UNAVAILABLE", "b-sla"))
    if updated_flag:
        badges.append(("attr", "UPDATED SINCE REVIEW", "b-updated"))
    return badges


# ---------------------------------------------------------------------------
# CSRF protection (loopback-appropriate)
# ---------------------------------------------------------------------------


def get_csrf_token():
    token = session.get("csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["csrf_token"] = token
    return token


def csrf_valid(token):
    expected = session.get("csrf_token")
    return bool(expected) and bool(token) and hmac.compare_digest(expected, token)


# ---------------------------------------------------------------------------
# Freshdesk ticket URLs (never followed in tests)
# ---------------------------------------------------------------------------


def ticket_url(ticket_id):
    return f"https://{FRESHDESK_DOMAIN}/a/tickets/{ticket_id}"


def fmt_due(due_str):
    if not due_str:
        return "—"
    try:
        dt = datetime.fromisoformat(due_str.replace("Z", "+00:00"))
        now = now_utc()
        delta = dt - now
        days = int(delta.total_seconds() // 86400)
        hours = int((delta.total_seconds() % 86400) // 3600)
        if delta.total_seconds() < 0:
            return f"<span style='color:red;font-weight:bold'>{abs(days)}d {abs(hours)}h OVERDUE</span>"
        return f"{days}d {hours}h left"
    except Exception:
        return due_str


# ---------------------------------------------------------------------------
# Closed-ticket housekeeping: offline-only search simulation
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class ClosedWindow:
    start: date
    end: date
    depth: int = 0


@dataclass
class ClosedRetrieval:
    tickets: list = field(default_factory=list)
    complete: bool = True
    warnings: list = field(default_factory=list)
    errors: list = field(default_factory=list)
    windows_planned: list = field(default_factory=list)
    windows_completed: list = field(default_factory=list)
    pages_requested: list = field(default_factory=list)
    reported_total_sum: int = 0
    unique_ticket_count: int = 0
    duplicate_count: int = 0
    date_range: tuple = None
    missing_tags_only: bool = True
    source_mode: str = "offline-synthetic"


def parse_closed_days(value) -> int:
    """Canonical positive integer day count; invalid values fail safely to 60."""
    try:
        if isinstance(value, bool) or not re.fullmatch(r"[0-9]+", str(value or "")):
            raise ValueError
        days = int(value)
        if not CLOSED_MIN_DAYS <= days <= CLOSED_MAX_DAYS:
            raise ValueError
        return days
    except (TypeError, ValueError):
        return CLOSED_DEFAULT_DAYS


def closed_filters_from_args(args):
    # A hidden form field supplies 0 when the checkbox is unchecked; use the
    # final value so checked 1 wins when Werkzeug exposes both values.
    values = args.getlist("missing_tags") if hasattr(args, "getlist") else []
    raw_missing = values[-1] if values else args.get("missing_tags", "1")
    # Photo/Video Review Scope defaults ON for Closed Ticket Housekeeping,
    # matching the main queue's default Review Scope. The closed page uses the
     # same canonical photo_video_only parameter and ticket-level matching
     # semantics (ticket_matches_photo_video); no second keyword list.
    pv_values = args.getlist("photo_video_only") if hasattr(args, "getlist") else []
    raw_pv = pv_values[-1] if pv_values else args.get("photo_video_only", "1")
    return {
        "days": parse_closed_days(args.get("days", CLOSED_DEFAULT_DAYS)),
        "missing_tags": parse_bool(raw_missing),
        "photo_video_only": parse_bool(raw_pv, default=True),
        "review_view": parse_review_view(args.get("review_view"), default="active"),
    }


def closed_query_string(start: date, end: date, missing_tags_only: bool) -> str:
    """Build a fixed, quoted Freshdesk search query from validated dates only."""
    if not isinstance(start, date) or not isinstance(end, date) or start > end:
        raise ValueError("Closed date range is invalid.")
    clauses = ["status:5"]
    if missing_tags_only:
        clauses.append("tag:null")
    clauses += [f"closed_at:>'{start.isoformat()}'", f"closed_at:<'{end.isoformat()}'"]
    return '"' + " AND ".join(clauses) + '"'


def closed_page_url(config) -> str:
    """Canonical /closed URL preserving days + missing_tags + review_view +
    photo_video_only.

    The review_view is a local dashboard filter only (it never appears in the
    Freshdesk search query); keeping it here makes every /closed URL
    bookmarkable and filter-preserving, exactly like /queue. The
    Photo/Video Review Scope is local-only as well and survives navigation.
    """
    missing = "1" if config.get("missing_tags") else "0"
    pv = "1" if config.get("photo_video_only", True) else "0"
    return (f"/closed?days={config.get('days', CLOSED_DEFAULT_DAYS)}"
            f"&missing_tags={missing}&photo_video_only={pv}"
            f"&review_view={config.get('review_view', 'active')}")


def closed_search_url_params(start: date, end: date, missing_tags_only: bool, page: int):
    if not isinstance(page, int) or not 1 <= page <= SEARCH_MAX_PAGE:
        raise ValueError("Search page must be between 1 and 10.")
    return urlencode({"query": closed_query_string(start, end, missing_tags_only), "page": page})


def split_closed_window(window: ClosedWindow):
    """Calendar midpoint split with one inclusive boundary overlap."""
    if window.start >= window.end:
        return None
    midpoint = window.start + timedelta(days=(window.end - window.start).days // 2)
    # Calendar ranges are partitioned, not overlapped, to guarantee a >300
    # two-day range can reduce. Deduplication still protects against an API
    # returning boundary duplicates in a future live transport.
    return (ClosedWindow(window.start, midpoint, window.depth + 1),
            ClosedWindow(midpoint + timedelta(days=1), window.end, window.depth + 1))


def _closed_at_date(ticket):
    raw = ticket.get("closed_at")
    if not isinstance(raw, str):
        return None
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def _synthetic_closed_tickets():
    """Synthetic-only fixture corpus. Large cases are generated deterministically."""
    base = [
        # Prompt 14: closed rows render queue-style Updated / Created columns,
        # so synthetic tickets carry updated_at/created_at (deterministic, from
        # closed_at; never derived from the wall clock). 810006/810007 keep the
        # missing/malformed closed-date coverage; 810006 also has NO
        # updated_at/created_at on purpose so the safe em-dash display path is
        # exercised end to end.
        {"id": 810001, "subject": "Synthetic closed untagged", "status": 5, "closed_at": "2026-08-04T09:00:00Z", "updated_at": "2026-08-04T10:00:00Z", "created_at": "2026-07-20T09:00:00Z", "tags": []},
        {"id": 810002, "subject": "Synthetic closed tagged", "status": 5, "closed_at": "2026-08-03T10:00:00Z", "updated_at": "2026-08-03T11:00:00Z", "created_at": "2026-07-15T09:00:00Z", "tags": ["parts"]},
        {"id": 810003, "subject": "Synthetic closed same timestamp low", "status": 5, "closed_at": "2026-08-02T10:00:00Z", "updated_at": "2026-08-02T11:00:00Z", "created_at": "2026-07-10T09:00:00Z", "tags": []},
        {"id": 810004, "subject": "Synthetic closed same timestamp high", "status": 5, "closed_at": "2026-08-02T10:00:00Z", "updated_at": "2026-08-02T11:30:00Z", "created_at": "2026-07-10T10:00:00Z", "tags": []},
        {"id": 810005, "subject": "Synthetic resolved excluded", "status": 4, "closed_at": "2026-08-01T10:00:00Z", "tags": []},
        {"id": 810006, "subject": "Synthetic missing date", "status": 5, "tags": []},
        {"id": 810007, "subject": "Synthetic malformed date", "status": 5, "closed_at": "not-a-date", "updated_at": "2026-08-03T12:00:00Z", "created_at": "2026-07-18T09:00:00Z", "tags": []},
        # Phase 4A: photo/video-subject closed tickets so the default Photo/Video
        # Review Scope has content on /closed. Subjects cover the canonical
        # keyword family (case-insensitive, word-boundary aware) and a couple of
        # non-photo/video subjects to prove the filter excludes them.
        {"id": 810010, "subject": "Customer sent photo of damage", "status": 5, "closed_at": "2026-08-04T08:00:00Z", "updated_at": "2026-08-04T09:00:00Z", "created_at": "2026-07-25T09:00:00Z", "tags": []},
        {"id": 810011, "subject": "Re: Photos of broken hinge", "status": 5, "closed_at": "2026-08-04T07:00:00Z", "updated_at": "2026-08-04T08:00:00Z", "created_at": "2026-07-25T09:00:00Z", "tags": ["parts"]},
        {"id": 810012, "subject": "Picture of scratched surface", "status": 5, "closed_at": "2026-08-03T08:00:00Z", "updated_at": "2026-08-03T09:00:00Z", "created_at": "2026-07-20T09:00:00Z", "tags": []},
        {"id": 810013, "subject": "Video of wobbling table leg", "status": 5, "closed_at": "2026-08-03T07:00:00Z", "updated_at": "2026-08-03T08:00:00Z", "created_at": "2026-07-20T09:00:00Z", "tags": []},
        {"id": 810014, "subject": "VID of damaged drawer", "status": 5, "closed_at": "2026-08-02T08:00:00Z", "updated_at": "2026-08-02T09:00:00Z", "created_at": "2026-07-15T09:00:00Z", "tags": []},
        {"id": 810015, "subject": "Pics attached for review", "status": 5, "closed_at": "2026-08-02T07:00:00Z", "updated_at": "2026-08-02T08:00:00Z", "created_at": "2026-07-15T09:00:00Z", "tags": []},
        {"id": 810016, "subject": "No photo here, just a question", "status": 5, "closed_at": "2026-08-01T08:00:00Z", "updated_at": "2026-08-01T09:00:00Z", "created_at": "2026-07-10T09:00:00Z", "tags": []},
        {"id": 810017, "subject": "Delivery schedule inquiry", "status": 5, "closed_at": "2026-08-01T07:00:00Z", "updated_at": "2026-08-01T08:00:00Z", "created_at": "2026-07-10T09:00:00Z", "tags": []},
        {"id": 810008, "subject": "Synthetic outside range", "status": 5, "closed_at": "2025-01-01T10:00:00Z", "updated_at": "2025-01-01T11:00:00Z", "created_at": "2024-12-01T09:00:00Z", "tags": []},
    ]
    # 301 tickets across two dates proves planner splitting without a 300-row file.
    for i in range(301):
        # Spread across 301 calendar dates so a >300 range splits cleanly;
        # custom fake transports cover the unsplittable single-day case.
        day = date(2025, 10, 8) + timedelta(days=i)
        base.append({"id": 820000 + i, "subject": f"Synthetic split ticket {i}", "status": 5,
                     "closed_at": f"{day.isoformat()}T12:00:00Z",
                     "updated_at": f"{day.isoformat()}T13:00:00Z",
                     "created_at": f"{(day - timedelta(days=30)).isoformat()}T09:00:00Z",
                     "tags": []})
    return base


def synthetic_closed_search_page(window: ClosedWindow, missing_tags_only: bool, page: int):
    """Injectable GET-like page reader. It has no network or credential path."""
    if page > SEARCH_MAX_PAGE:
        raise ValueError("Synthetic transport refuses page 11.")
    matches = []
    for ticket in _synthetic_closed_tickets():
        closed = _closed_at_date(ticket)
        if ticket.get("status") != CLOSED_STATUS or closed is None:
            continue
        if missing_tags_only and ticket.get("tags"):
            continue
        if window.start <= closed <= window.end:
            matches.append(ticket)
    # Deliberate duplicate across pages/windows to exercise deduplication.
    matches.sort(key=lambda t: t["id"])
    start = (page - 1) * SEARCH_PAGE_SIZE
    return {"total": len(matches), "results": matches[start:start + SEARCH_PAGE_SIZE]}


def retrieve_closed_tickets(start: date, end: date, missing_tags_only: bool,
                             search_page: Callable = synthetic_closed_search_page):
    """Deterministic search planning. The transport returns dict(total, results)."""
    result = ClosedRetrieval(date_range=(start, end), missing_tags_only=missing_tags_only)
    pending = [ClosedWindow(start, end)]
    seen = {}
    while pending:
        if len(result.windows_planned) >= CLOSED_MAX_WINDOWS:
            result.complete = False; result.errors.append("Maximum date-window count reached."); break
        window = pending.pop(0); result.windows_planned.append(window)
        try:
            first = search_page(window, missing_tags_only, 1)
            if not isinstance(first, dict) or not isinstance(first.get("total"), int) or not isinstance(first.get("results"), list):
                raise ValueError("Malformed search response: expected integer total and results list.")
            total = first["total"]
            if total < 0: raise ValueError("Malformed search response: negative total.")
            result.reported_total_sum += total
            if total > SEARCH_MAX_RESULTS:
                children = split_closed_window(window)
                if children is None:
                    result.complete = False; result.errors.append("More than 300 matching tickets were closed on one date. This range cannot be fully retrieved through the Search Tickets page limit.")
                    continue
                if window.depth >= CLOSED_MAX_SPLIT_DEPTH:
                    result.complete = False; result.errors.append("Maximum date-window split depth reached."); continue
                pending[0:0] = children
                continue
            pages = ceil(total / SEARCH_PAGE_SIZE)
            responses = [first]
            result.pages_requested.append((window, 1))
            for page in range(2, pages + 1):
                result.pages_requested.append((window, page))
                response = search_page(window, missing_tags_only, page)
                if not isinstance(response, dict) or response.get("total") != total or not isinstance(response.get("results"), list):
                    raise ValueError("Malformed or inconsistent search page response.")
                responses.append(response)
            received = [t for response in responses for t in response["results"]]
            if len(received) < total:
                raise ValueError("Search page ended before the reported total was retrieved.")
            for ticket in received:
                if not isinstance(ticket, dict) or not isinstance(ticket.get("id"), int):
                    raise ValueError("Malformed ticket in search response.")
                ticket_date = _closed_at_date(ticket)
                if ticket.get("status") != CLOSED_STATUS or ticket_date is None or not window.start <= ticket_date <= window.end:
                    result.complete = False; result.warnings.append("Ignored malformed or out-of-contract ticket from synthetic response."); continue
                if missing_tags_only and ticket.get("tags"):
                    continue
                if ticket["id"] in seen: result.duplicate_count += 1
                seen[ticket["id"]] = ticket
            result.windows_completed.append(window)
        except Exception as exc:
            result.complete = False
            result.errors.append(str(exc))
    result.tickets = sorted(seen.values(), key=lambda t: (_closed_at_date(t), t["id"]), reverse=True)
    result.unique_ticket_count = len(result.tickets)
    return result


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------


def _queue_error_page(message, offline):
    return _queue_render(
        tickets=[], total=0, offline=offline, cache_age=None,
        error=message, config=dict(DEFAULT_FILTERS), all_categories_off=False,
    )


def closed_display(value: str) -> str:
    """ISO-8601 closed_at -> compact 'YYYY-MM-DD HH:MM' (queue date-column style).

    Presentation-only: the raw ISO value stays available in the ticket dict
    (``closed_at``); this helper just makes the Closed column compact and
    scannable the way /queue renders its date columns. Missing, empty, or
    malformed values render as an em dash (the /queue date convention) rather
    than a raw string, and never crash.
    """
    if not value:
        return "—"
    s = value.replace("T", " ").replace("Z", "").strip()
    if len(s) < 10 or s[4] != "-" or s[7] != "-":
        return "—"
    return s[:16]


_EASTERN = ZoneInfo("America/New_York")


def format_eastern_timestamp(value) -> str:
    """ISO-8601 UTC timestamp -> Eastern local time as 'M/D/YY h:mm AM/PM TZ'.

    Presentation-only: the raw ``value`` is never mutated. Uses
    ``America/New_York`` so the abbreviation (EDT/EST) tracks DST
    automatically. Missing, empty, malformed, or timezone-less values render as
    an em dash and never crash."""
    parsed = parse_dt(value)
    if parsed is None:
        return "—"
    local = parsed.astimezone(_EASTERN)
    # M/D/YY h:mm AM/PM with no leading zeros on month/day/hour.
    return f"{local.month}/{local.day}/{local.year % 100:02d} {local.strftime('%I:%M %p %Z').lstrip('0')}"


CLOSED_CACHE_MAX_AGE_SECONDS = 24 * 60 * 60


def closed_live_result(config, now=None):
    """Build a ClosedRetrieval from the LIVE cache file only.

    This never performs a request and never reads the API key: it is pure
    presentation over ``cache/closed_tickets.json``, which only an explicit
    refresh can write. Membership uses the exact half-open UTC boundaries
    ``start <= stats.closed_at < end`` — the same window the refresh job used.
    """
    now = now or now_utc()
    start, end = closed_live.utc_window(config["days"], now)
    result = ClosedRetrieval(
        date_range=(start.date(), (end - timedelta(days=1)).date()),
        missing_tags_only=config["missing_tags"],
    )
    result.source_mode = "live-cache"
    blob = closed_live.load_cache()
    if blob is None:
        result.complete = False
        result.warnings.append(
            "No live closed data has been retrieved yet. Use Refresh Closed Data to fetch it."
        )
        return result

    kept, missing_closed_at, missing_tags_unknown = [], 0, 0
    for ticket in blob["tickets"]:
        if ticket.get("status") != CLOSED_STATUS:
            continue
        raw = closed_live.cache_closed_at(ticket)
        closed_dt = parse_dt(raw)
        if closed_dt is None:
            missing_closed_at += 1
            continue
        if not (start <= closed_dt < end):
            continue
        tags = ticket.get("tags")
        if not isinstance(tags, list):
            missing_tags_unknown += 1
            continue
        if config["missing_tags"] and tags:
            continue
        kept.append(dict(ticket, closed_at=raw, tags=tags))

    kept.sort(key=lambda t: (t.get("closed_at") or "", t["id"]), reverse=True)
    result.tickets = kept
    result.unique_ticket_count = len(kept)

    # Coverage reporting: the operator must be able to see when the cached
    # window is narrower or older than what they are asking to view.
    coverage = blob.get("coverage") if isinstance(blob.get("coverage"), dict) else {}
    cached_days = blob.get("days")
    if isinstance(cached_days, int) and cached_days < config["days"]:
        result.complete = False
        result.warnings.append(
            f"Coverage warning: cached data only covers the last {cached_days} days, "
            f"but {config['days']} days are being displayed. Refresh to widen coverage."
        )
    if coverage.get("next_page_existed_at_cap"):
        result.complete = False
        result.warnings.append(
            "Coverage warning: the last refresh hit the page ceiling, so older tickets may be missing."
        )
    if missing_closed_at:
        result.warnings.append(
            f"Coverage warning: {missing_closed_at} cached tickets have no usable closed_at and were excluded."
        )
    if missing_tags_unknown:
        result.warnings.append(
            f"Coverage warning: {missing_tags_unknown} cached tickets have unknown tag data and were excluded."
        )
    fetched = parse_dt(blob.get("fetched_at"))
    if fetched is None:
        result.warnings.append("Coverage warning: cached data has no usable retrieval timestamp.")
    elif (now - fetched).total_seconds() > CLOSED_CACHE_MAX_AGE_SECONDS:
        result.warnings.append(
            f"Coverage warning: cached data was retrieved {fetched.isoformat()} and may be stale."
        )
    result.pages_requested = [None] * int(coverage.get("pages_completed") or 0)
    result.windows_planned = [ClosedWindow(start.date(), (end - timedelta(days=1)).date())]
    result.windows_completed = list(result.windows_planned)
    return result


@app.route("/closed")
def closed_housekeeping():
    """Closed-ticket page.

    Offline mode renders the untouched synthetic fixture corpus. Live mode
    renders ONLY the explicitly refreshed cache file — the page itself never
    triggers a request and never reads the API key.
    """
    config = closed_filters_from_args(request.args)
    offline = is_offline()
    # Local date is intentionally used for display range. Search predicates use
    # explicit YYYY-MM-DD calendar dates; see the UTC caveat in the contract doc.
    end = now_utc().date()
    start = end - timedelta(days=config["days"] - 1)
    try:
        if offline:
            result = retrieve_closed_tickets(start, end, config["missing_tags"])
        else:
            result = closed_live_result(config)
    except Exception:
        return _closed_render(result=None, config=config, offline=offline,
                              error="Closed retrieval failed safely. No live fallback was attempted."), 500

    # Closed-housekeeping local review state (Prompt 12) — separate namespace
    # from /queue (closed_review_state table), applied after retrieval so the
    # retrieval metadata (windows/pages/dupes/sorting) stays intact.
    #
    # Photo/Video Review Scope: an additional local-only layer applied alongside
    # the existing review_view filter. Uses the SAME ticket_matches_photo_video
    # matcher as the main queue — subject or valid tag, word-boundary aware,
    # case-insensitive. Defaults ON (see closed_filters_from_args). OFF shows
    # the full closed population that satisfies the remaining filters. Never
    # triggers retrieval, never writes cache, never changes review state.
    closed_rows = load_closed_review_rows()
    closed_last_opened = closed_last_opened_ticket_id()
    reviewed = []
    for t in result.tickets:
        tid = t["id"]
        state_row = closed_rows.get(tid)
        result_state = state_row.get("review_result", "Unreviewed") if state_row else "Unreviewed"
        if not closed_review_view_includes(state_row, config["review_view"]):
            continue
        if config.get("photo_video_only") and not ticket_matches_photo_video(t):
            continue
        row_class = REVIEW_CLASS.get(result_state, "rv-unreviewed")
        is_last_opened = closed_last_opened is not None and tid == closed_last_opened
        if is_last_opened:
            row_class += " rv-last-opened"
        badges = []
        badge_text = "OPENED / IN REVIEW" if result_state == "Opened / In Review" else result_state
        badges.append(("review", badge_text, "b-review " + REVIEW_CLASS.get(result_state, "rv-unreviewed")))
        if is_last_opened:
            badges.append(("focus", "LAST OPENED", "b-last-opened"))
        if not t.get("tags"):
            badges.append(("attr", "MISSING TAGS", "b-missing"))
        reviewed.append({
            "id": tid,
            "url": ticket_url(tid),
            "subject": t.get("subject", ""),
            "status": t.get("status"),
            "status_label": STATUS_LABELS.get(t.get("status"), f"Status {t.get('status')}"),
            "closed_at": t.get("closed_at", ""),
            "closed_display": closed_display(t.get("closed_at", "")),
            "updated_display": format_eastern_timestamp(t.get("updated_at")),
            "created_display": (t.get("created_at") or "")[:10] or "—",
            "tags": t.get("tags") or [],
            "result": result_state,
            "row_class": row_class,
            "last_opened": is_last_opened,
            "badges": badges,
        })
    result.tickets = reviewed
    flash_msg = session.pop("flash", None)
    return _closed_render(
        result=result, config=config, error=None, view_count=len(reviewed),
        csrf_token=get_csrf_token(), flash=flash_msg,
        review_states=REVIEW_STATES, offline=offline,
        closed_last_opened=closed_last_opened,
        last_opened_rendered=any(r["last_opened"] for r in reviewed),
    )


@app.route("/closed/api/review", methods=["POST"])
def closed_review_api():
    """Save a local closed-housekeeping review result (form POST). Redirects
    back to the exact /closed view the user was on, preserving days,
    missing_tags, and review_view. Never changes Freshdesk."""
    config = closed_filters_from_args(request.form)

    def back_to_closed(msg, ok):
        session["flash"] = ("ok" if ok else "err", msg)
        return redirect(closed_page_url(config), code=303)

    if not csrf_valid(request.form.get("csrf_token")):
        return back_to_closed("Review not saved: invalid security token. Reload the page and try again.", False)

    raw_ticket_id = request.form.get("ticket_id")
    if raw_ticket_id is None or str(raw_ticket_id).strip() == "":
        return back_to_closed("Review not saved: missing ticket ID.", False)
    try:
        ticket_id = int(raw_ticket_id)
    except (TypeError, ValueError):
        return back_to_closed("Review not saved: invalid ticket ID.", False)

    result = request.form.get("review_result")
    if result not in REVIEW_STATES:
        return back_to_closed("Review not saved: unknown review result.", False)

    # The ticket must belong to the closed corpus. No live lookup is ever
    # attempted (offline mode), so an unknown id is rejected without network.
    if not closed_ticket_known(ticket_id):
        return back_to_closed(f"Review not saved: unknown ticket #{ticket_id}.", False)

    try:
        set_closed_review_result(ticket_id, result)
    except Exception as e:
        return back_to_closed(f"Review not saved: database error ({e}).", False)

    return back_to_closed(f"Review saved for #{ticket_id}: {result}.", True)


@app.route("/closed/api/refresh", methods=["POST"])
def closed_refresh_start():
    """Start one explicit read-only Freshdesk list-ticket refresh."""
    if not csrf_valid(request.form.get("csrf_token")):
        return jsonify({"ok": False, "message": "Invalid security token."}), 403
    if is_offline():
        return jsonify({"ok": False, "message": "Offline mode: live refresh is disabled."}), 409
    config = closed_filters_from_args(request.form)
    try:
        api_key = load_api_key()
    except Exception:
        return jsonify({"ok": False, "message": "Freshdesk credential is unavailable."}), 503
    started, message = closed_live.JOB.start(days=config["days"], api_key=api_key, now=now_utc())
    status = closed_live.JOB.status()
    status.update({"ok": started, "message": message})
    return jsonify(status), (202 if started else 409)


@app.route("/closed/api/refresh/status")
def closed_refresh_status():
    """Safe aggregate status only; ticket payloads are never returned."""
    return jsonify(closed_live.JOB.status())


@app.route("/closed/api/refresh/cancel", methods=["POST"])
def closed_refresh_cancel():
    if not csrf_valid(request.form.get("csrf_token")):
        return jsonify({"ok": False, "message": "Invalid security token."}), 403
    cancelled = closed_live.JOB.cancel()
    return jsonify({
        "ok": cancelled,
        "message": "Cancellation requested." if cancelled else "No refresh is running.",
    }), (202 if cancelled else 409)


@app.route("/closed/api/opened", methods=["POST"])
def closed_opened_api():
    """Record that a closed ticket link was opened (JSON). Same contract as the
    queue's opened endpoint: the click handler never prevents the new tab, so
    the Freshdesk ticket always opens even if recording fails; the UI only
    claims the save when this endpoint returns ok. Local state only."""
    if not is_offline():
        return jsonify({"ok": False, "error": "offline-only endpoint"}), 503
    token = request.headers.get("X-CSRF-Token") or (request.get_json(silent=True) or {}).get("csrf_token")
    if not csrf_valid(token):
        return jsonify({"ok": False, "error": "invalid security token"}), 403

    body = request.get_json(silent=True) or {}
    raw_ticket_id = body.get("ticket_id")
    if raw_ticket_id is None:
        return jsonify({"ok": False, "error": "missing ticket_id"}), 400
    try:
        ticket_id = int(raw_ticket_id)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "invalid ticket_id"}), 400

    if not closed_ticket_known(ticket_id):
        return jsonify({"ok": False, "error": "unknown ticket id"}), 404

    try:
        result = mark_closed_opened(ticket_id)
    except Exception as e:
        return jsonify({"ok": False, "error": f"database error: {e}"}), 500

    return jsonify({
        "ok": True,
        "review_result": result,
        "last_opened_id": closed_last_opened_ticket_id(),
    })


def closed_ticket_known(ticket_id):
    """True when the id is present in the active offline fixture or live cache."""
    if is_offline():
        return any(
            t.get("status") == CLOSED_STATUS and t.get("id") == ticket_id
            for t in _synthetic_closed_tickets()
        )
    payload = closed_live.load_cache()
    if not payload:
        return False
    return any(t.get("id") == ticket_id for t in payload.get("tickets", []))


def build_current_queue_view(raw, config):
    """Build the exact ticket rows represented by the queue dashboard.

    This helper is render-only: it reads the supplied cache tickets and local
    review state, and never refreshes, opens, or mutates anything.
    """
    all_categories_off = True
    closed_mode = config["mode"] == "closed"
    show_all_cached = False
    if closed_mode:
        tickets = [t for t in raw if t.get("status") == CLOSED_STATUS or str(t.get("status")) == str(CLOSED_STATUS)]
        if config["missing_tags"]:
            tickets = [t for t in tickets if not isinstance(t.get("tags"), list) or not t.get("tags")]
        if config["photo_video_only"]:
            tickets = [t for t in tickets if ticket_matches_photo_video(t)]
    else:
        normal_config = dict(config, overdue=False, responded=False, waiting=False)
        tickets = apply_queue_filters(raw, normal_config)
        show_all_cached = not any((config["photo_video_only"], config["hide_reviewed_tags"], config["missing_tags"]))
        if not show_all_cached:
            tickets = [t for t in tickets if t.get("status") != CLOSED_STATUS and str(t.get("status")) != str(CLOSED_STATUS)]

    queue_scope_counts = {"main": 0, "triage": 0}
    if not closed_mode:
        classified_tickets = []
        for ticket in tickets:
            scope = "main" if is_main_queue_ticket(ticket) else "triage"
            queue_scope_counts[scope] += 1
            if scope == config.get("queue_scope", "main"):
                classified_tickets.append(ticket)
        tickets = classified_tickets
        # Offline fixtures predate the Freshdesk dashboard metadata fields. Keep
        # fixture-based regression tests exercising the legacy review workflow;
        # live cached preview data always uses the strict Main/Triage split.
        if is_offline() and os.environ.get("FRESHDESK_OFFLINE_CACHE", "").strip().lower() not in ("1", "true", "yes"):
            tickets = apply_queue_filters(raw, normal_config)
            if not show_all_cached:
                tickets = [ticket for ticket in tickets if main_queue_status(ticket) != CLOSED_STATUS]
            queue_scope_counts = {"main": len(tickets), "triage": 0}

    state_rows = load_review_rows()
    last_opened_id = last_opened_ticket_id()
    rows = []
    workflow_counts = {tab: 0 for tab in WORKFLOW_TABS}
    ordered_tickets = sorted(tickets, key=lambda x: (not updated_since_review(x, state_rows.get(x.get("id"))), x.get("id") is None, x.get("id") or 0))
    for t in ordered_tickets:
        tid = t["id"]
        state_row = state_rows.get(tid)
        updated_flag = updated_since_review(t, state_row)
        destination = workflow_destination(state_row.get("review_result", "Unreviewed") if state_row else "Unreviewed", updated_flag)
        workflow_counts[destination] += 1
        if not show_all_cached and destination != config["workflow_tab"]:
            continue
        sid = t.get("status")
        pid = t.get("priority", 0)
        due = t.get("due_by") or t.get("fr_due_by")
        updated_at = t.get("updated_at")
        row_class = REVIEW_CLASS.get(state_row.get("review_result", "Unreviewed") if state_row else "Unreviewed", "rv-unreviewed")
        is_last_opened = last_opened_id is not None and tid == last_opened_id
        if is_last_opened:
            row_class += " rv-last-opened"
        rows.append({
            "id": tid, "url": ticket_url(tid), "subject": t.get("subject", ""),
            "status_label": STATUS_LABELS.get(sid, f"Status {sid}"),
            "priority_label": PRIORITY_LABELS.get(pid, f"P{pid}"),
            "updated_display": format_eastern_timestamp(updated_at),
            "created_display": (t.get("created_at") or "")[:10] or "—",
            "tags": (t.get("tags") or []) if isinstance(t.get("tags"), list) else [],
            "result": state_row.get("review_result", "Unreviewed") if state_row else "Unreviewed",
            "row_class": row_class, "last_opened": is_last_opened,
            "updated_flag": updated_flag,
            "triage_reasons": [] if closed_mode or is_main_queue_ticket(t) else main_queue_triage_reasons(t),
            "badges": ticket_badges(t, state_row, updated_flag),
            "can_acknowledge": bool(state_row and state_row.get("review_result") in REVIEWED_STATES and updated_flag),
        })
    return rows, workflow_counts, queue_scope_counts, all_categories_off, last_opened_id, show_all_cached


@app.route("/queue")
def queue():
    """GET /queue — RENDER ONLY. Never fetches from Freshdesk.

    Renders filter controls, uses whatever is already stored in the LIVE queue
    cache (live mode) or the offline fixtures (offline mode), applies local
    filters and local SQLite review state. Missing/stale/absent cache in live
    mode is a neutral "click Refresh Tickets" state — it is NOT auto-refreshed.
    """
    config = filters_from_args(request.args)
    offline = is_offline()

    # Missing-key warning so the user notices before a blank page. Skipped in
    # offline mode — offline mode works without a key and never reads it. The
    # live page still renders the filter controls and the neutral Refresh Tickets cue so
    # the operator always has a valid UI state; Refresh itself explains the
    # missing key when clicked.
    missing_key_msg = None
    if not offline and not load_api_key():
        missing_key_msg = (
            "No Freshdesk API key found. Set FRESHDESK_API_KEY env var or write it to "
            "~/.config/furtouch/freshdesk_api_key (chmod 600). Refresh Tickets will not work "
            "until a key is available."
        )

    # RENDER ONLY — never fetches. OfflineDataError is the only failure path;
    # live mode always renders successfully (possibly an empty pool).
    try:
        raw, cache_age = get_ticket_pool()
        blob = None if offline else load_live_queue_cache()
        cached_days = blob.get("days") if blob and isinstance(blob.get("days"), int) else None
    except OfflineDataError as e:
        return _queue_error_page(str(e), offline)

    rows, workflow_counts, queue_scope_counts, all_categories_off, last_opened_id, show_all_cached = build_current_queue_view(raw, config)
    flash_msg = session.pop("flash", None)
    return _queue_render(
        tickets=rows, total=len(rows), error=missing_key_msg,
        workflow_counts=workflow_counts, queue_scope_counts=queue_scope_counts,
        offline=offline, cache_age=cache_age,
        config=config, csrf_token=get_csrf_token(), flash=flash_msg,
        all_categories_off=all_categories_off, last_opened_id=last_opened_id,
        last_opened_rendered=any(r["last_opened"] for r in rows),
        live_cache_missing=(None is cache_age and not offline and len(raw) == 0),
        cached_days=cached_days, cached_ticket_count=len(raw),
        last_refresh_display=_last_refresh_display(cache_age, offline),
        last_refresh_mode_display=((blob or {}).get("last_refresh_mode") or "No baseline").replace("baseline", "Baseline").replace("incremental", "Incremental").replace("reconcile", "Reconcile"),
        cache_coverage_display=_cache_coverage_display(cached_days, config["days"]),
        cache_coverage_warning=(cached_days is None or (isinstance(cached_days, int) and config["days"] > cached_days)),
        auto_refresh=auto_refresh_status(),
    )

    # Default state: Review Scope controls are ON (photo/video subjects only +
    # reviewed/closed tag exclusions), all manual controls are OFF. The visible
    # "Show All Cached Tickets" control turns the scope controls and every
    # manual control off for a complete-cache view.
    # ``all_categories_off`` is retained as template/context compatibility but
    # no longer suppresses rows.
    all_categories_off = True  # retained template compatibility; obsolete selectors are inert.
    closed_mode = config["mode"] == "closed"
    show_all_cached = False
    if closed_mode:
        # The master 60-day queue cache is the sole Closed Housekeeping source.
        # No closed_at window and no closed_tickets.json lookup are involved.
        tickets = [t for t in raw if t.get("status") == CLOSED_STATUS or str(t.get("status")) == str(CLOSED_STATUS)]
        if config["missing_tags"]:
            tickets = [t for t in tickets if not isinstance(t.get("tags"), list) or not t.get("tags")]
        if config["photo_video_only"]:
            tickets = [t for t in tickets if ticket_matches_photo_video(t)]
    else:
        # Obsolete legacy URL fields are retained for bookmark serialization,
        # but must not reintroduce their former queue restrictions.
        normal_config = dict(config, overdue=False, responded=False, waiting=False)
        tickets = apply_queue_filters(raw, normal_config)
        # The explicit "Show All Cached Tickets" diagnostic escape hatch is
        # every remaining local scope/filter control OFF; it retains full-cache
        # semantics including Closed tickets. Otherwise Normal Review excludes
        # actual Freshdesk Closed status.
        show_all_cached = not any((config["photo_video_only"], config["hide_reviewed_tags"], config["missing_tags"]))
        if not show_all_cached:
            tickets = [t for t in tickets if t.get("status") != CLOSED_STATUS and str(t.get("status")) != str(CLOSED_STATUS)]

    state_rows = load_review_rows()
    last_opened_id = last_opened_ticket_id()  # focus state, independent of filters

    # Per-ticket workflow decision + row data. Closed tickets are omitted from
    # normal workflow tabs, while explicit Show All remains diagnostic.
    #
    # WORKFLOW ROUTING IS AUTHORITATIVE. The legacy review_view URL parameter
    # is still parsed and preserved for backward compatibility (it is emitted
    # by the filter form, preset links, and the refresh-completion redirect),
    # but it must NEVER override the active workflow-tab routing. Without
    # this guarantee, a post-refresh URL like "/queue?review_view=all" (no
    # workflow_tab) would display every reviewed ticket in the Main Queue
    # table while the Main Queue tab count correctly read 0 — the exact
    # production bug this guard prevents.
    rows = []
    workflow_counts = {tab: 0 for tab in WORKFLOW_TABS}
    ordered_tickets = sorted(tickets, key=lambda x: (not updated_since_review(x, state_rows.get(x.get("id"))), x.get("id") is None, x.get("id") or 0))
    for t in ordered_tickets:
        tid = t["id"]
        state_row = state_rows.get(tid)
        updated_flag = updated_since_review(t, state_row)
        destination = workflow_destination(state_row.get("review_result", "Unreviewed") if state_row else "Unreviewed", updated_flag)
        workflow_counts[destination] += 1
        if not show_all_cached and destination != config["workflow_tab"]:
            continue
        sid = t.get("status")
        pid = t.get("priority", 0)
        due = t.get("due_by") or t.get("fr_due_by")
        updated_at = t.get("updated_at")
        row_class = REVIEW_CLASS.get(
            state_row.get("review_result", "Unreviewed") if state_row else "Unreviewed",
            "rv-unreviewed",
        )
        is_last_opened = last_opened_id is not None and tid == last_opened_id
        if is_last_opened:
            row_class += " rv-last-opened"  # layered on top of the review class
        rows.append({
            "id": tid,
            "url": ticket_url(tid),
            "subject": t.get("subject", ""),
            "status_label": STATUS_LABELS.get(sid, f"Status {sid}"),
            "priority_label": PRIORITY_LABELS.get(pid, f"P{pid}"),
            "due_display": fmt_due(due),
            "updated_display": format_eastern_timestamp(updated_at),
            "created_display": (t.get("created_at") or "")[:10] or "—",
            "tags": (t.get("tags") or []) if isinstance(t.get("tags"), list) else [],
            "type": t.get("type"),
            "result": state_row.get("review_result", "Unreviewed") if state_row else "Unreviewed",
            "row_class": row_class,
            "last_opened": is_last_opened,
             "updated_flag": updated_flag,
             "badges": ticket_badges(t, state_row, updated_flag),
             "can_acknowledge": bool(state_row and state_row.get("review_result") in REVIEWED_STATES and updated_flag),
        })

    flash_msg = session.pop("flash", None)

    return _queue_render(
        tickets=rows, total=len(rows), error=missing_key_msg,
        workflow_counts=workflow_counts,
        offline=offline, cache_age=cache_age, config=config,
        csrf_token=get_csrf_token(), flash=flash_msg,
        all_categories_off=all_categories_off,
        last_opened_id=last_opened_id,
        last_opened_rendered=any(r["last_opened"] for r in rows),
         live_cache_missing=(None is cache_age and not offline and len(raw) == 0),
         cached_days=cached_days,
         cached_ticket_count=len(raw),
         last_refresh_mode_display=((blob or {}).get("last_refresh_mode") or "No baseline").replace("baseline", "Baseline").replace("incremental", "Incremental").replace("reconcile", "Reconcile"),
         last_refresh_display=_last_refresh_display(cache_age, offline),
          cache_coverage_display=_cache_coverage_display(cached_days, config["days"]),
          cache_coverage_warning=(cached_days is None or (isinstance(cached_days, int) and config["days"] > cached_days)),
          auto_refresh=auto_refresh_status(),
     )



@app.route("/queue/export.xlsx")
def queue_export_xlsx():
    """Download the exact current queue view as a read-only XLSX workbook."""
    config = filters_from_args(request.args)
    try:
        raw, _cache_age = get_ticket_pool()
    except OfflineDataError as e:
        return _queue_error_page(str(e), is_offline())
    rows, _counts, _scope_counts, _all_categories_off, _last_opened_id, _show_all = build_current_queue_view(raw, config)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Current View"
    headers = ["Ticket #", "Subject", "Status", "Priority", "Updated", "Created", "Tags", "Review Result"]
    if config["mode"] == "normal" and config.get("queue_scope", "main") == "triage":
        headers.append("Why here?")
    headers.append("Freshdesk URL")
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        values = [
            row["id"], row["subject"], row["status_label"], row["priority_label"],
            row["updated_display"], row["created_display"], ", ".join(str(tag) for tag in row["tags"]),
            row["result"],
        ]
        if config["mode"] == "normal" and config.get("queue_scope", "main") == "triage":
            values.append("; ".join(row["triage_reasons"]))
        values.append(row["url"])
        sheet.append(values)
        url_cell = sheet.cell(sheet.max_row, len(headers))
        url_cell.hyperlink = row["url"]
        url_cell.style = "Hyperlink"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(sheet.max_row, 1)}"
    widths = {"A": 11, "B": 42, "C": 22, "D": 12, "E": 24, "F": 14, "G": 35, "H": 24, "I": 55}
    if config["mode"] == "normal" and config.get("queue_scope", "main") == "triage":
        widths["I"] = 42
        widths["J"] = 55
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[6].alignment = Alignment(wrap_text=True, vertical="top")
        row[len(headers) - 1].alignment = Alignment(wrap_text=True, vertical="top")

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"freshdesk-current-view-{date.today().isoformat()}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _start_normal_queue_refresh():
    """Start the existing normal incremental refresh path for a user or scheduler."""
    try:
        api_key = load_api_key()
    except Exception:
        api_key = ""
    if not api_key:
        return False, "No Freshdesk API key is available."
    old_blob = load_live_queue_cache()
    attempt_started_at = now_utc().astimezone(timezone.utc).replace(microsecond=0)
    if old_blob is None and os.path.exists(LIVE_QUEUE_CACHE_FILE):
        return False, "Existing queue cache is malformed; refresh was not started."
    return queue_live.JOB.start(
        days=DAYS_DEFAULT, api_key=api_key,
        retrieve=fetch_live_queue,
        save=save_live_queue_cache,
        finalize=lambda incoming_tickets, progress_callback=None, cancel_callback=None,
                        attempt_started_at=None: _reconcile_queue_refresh(
            old_blob, incoming_tickets, progress_callback=progress_callback,
            cancel_callback=cancel_callback, attempt_started_at=attempt_started_at,
            current_agent_id_fetcher=lambda: fetch_current_agent_id(api_key),
        ),
        plan=lambda requested_days, job_attempt_started_at: queue_refresh_plan(
            old_blob, requested_days, job_attempt_started_at),
        attempt_started_at=attempt_started_at, mode="normal",
    )


@app.route("/queue/api/refresh", methods=["POST"])
def queue_refresh_start():
    """Start one finite normal or explicit reconcile refresh."""
    mode = request.form.get("mode", "normal").strip().lower()
    if mode not in {"normal", "reconcile"}:
        return jsonify({"ok": False, "message": "Invalid refresh mode."}), 400
    raw_days = request.form.get("days")
    if mode == "reconcile":
        if raw_days is None or not str(raw_days).strip().isdigit() or not (DAYS_MIN <= int(str(raw_days).strip()) <= DAYS_MAX):
            return jsonify({"ok": False, "message": f"Reconcile range must be an integer from {DAYS_MIN} to {DAYS_MAX} days."}), 400
        days = int(str(raw_days).strip())
    else:
        days = DAYS_DEFAULT
    if not csrf_valid(request.form.get("csrf_token")):
        return jsonify({"ok": False, "message": "Invalid security token."}), 403
    if is_offline():
        return jsonify({"ok": False, "message": "Offline mode: live refresh is disabled."}), 409
    if mode == "normal":
        started, message = _start_normal_queue_refresh()
    else:
        try:
            api_key = load_api_key()
        except Exception:
            api_key = ""
        if not api_key:
            return jsonify({"ok": False, "message": "No Freshdesk API key is available."}), 503
        old_blob = load_live_queue_cache()
        attempt_started_at = now_utc().astimezone(timezone.utc).replace(microsecond=0)
        if old_blob is None and os.path.exists(LIVE_QUEUE_CACHE_FILE):
            return jsonify({"ok": False, "message": "Existing queue cache is malformed; refresh was not started."}), 409
        started, message = queue_live.JOB.start(
            days=days, api_key=api_key,
            retrieve=fetch_live_queue,
            save=save_live_queue_cache,
            finalize=lambda incoming_tickets, progress_callback=None, cancel_callback=None,
                            attempt_started_at=None: _reconcile_queue_refresh(
                old_blob, incoming_tickets, progress_callback=progress_callback,
                cancel_callback=cancel_callback, attempt_started_at=attempt_started_at,
                current_agent_id_fetcher=lambda: fetch_current_agent_id(api_key),
            ),
            plan=lambda requested_days, job_attempt_started_at, requested_mode="normal": queue_refresh_plan(
                old_blob, requested_days, job_attempt_started_at, requested_mode),
            attempt_started_at=attempt_started_at, mode=mode,
        )
    if started:
        reset_auto_refresh_countdown()
    status = queue_live.JOB.status()
    status.update({"ok": started, "message": message})
    return jsonify(status), (202 if started else 409)


@app.route("/queue/api/refresh/status")
def queue_refresh_status():
    """Return local in-memory queue refresh and auto-refresh state only."""
    status = queue_live.JOB.status()
    status["auto_refresh"] = auto_refresh_status()
    return jsonify(status)


@app.route("/queue/api/refresh/cancel", methods=["POST"])
def queue_refresh_cancel():
    if not csrf_valid(request.form.get("csrf_token")):
        return jsonify({"ok": False, "message": "Invalid security token."}), 403
    cancelled = queue_live.JOB.cancel()
    return jsonify({"ok": cancelled, "message": "Cancel requested." if cancelled else "No queue refresh is running."})


@app.route("/queue/api/review", methods=["POST"])
def review_api():
    """Save a local review result (form POST). Redirects back to the exact
    queue view the user was on, preserving every filter parameter."""
    config = filters_from_args(request.form)

    def back_to_queue(msg, ok):
        session["flash"] = ("ok" if ok else "err", msg)
        return redirect(f"/queue?{filter_query_string(config)}", code=303)

    if not csrf_valid(request.form.get("csrf_token")):
        return back_to_queue("Review not saved: invalid security token. Reload the page and try again.", False)

    raw_ticket_id = request.form.get("ticket_id")
    if raw_ticket_id is None or str(raw_ticket_id).strip() == "":
        return back_to_queue("Review not saved: missing ticket ID.", False)
    try:
        ticket_id = int(raw_ticket_id)
    except (TypeError, ValueError):
        return back_to_queue("Review not saved: invalid ticket ID.", False)

    result = request.form.get("review_result")
    if result not in REVIEW_STATES:
        return back_to_queue("Review not saved: unknown review result.", False)

    # Snapshot the ticket's current updated_at for Reviewed states. The lookup
    # reads the same pool the dashboard uses (cache-first; offline = fixtures,
    # so no HTTP is ever triggered by a review update).
    reviewed_updated_at = None
    try:
        pool, _ = get_ticket_pool()
        match = next((t for t in pool if t.get("id") == ticket_id), None)
    except Exception:
        match = None
    if match is None:
        return back_to_queue(f"Review not saved: unknown ticket #{ticket_id}.", False)
    if result in REVIEWED_STATES:
        reviewed_updated_at = match.get("updated_at")

    try:
        set_review_result(ticket_id, result, reviewed_updated_at)
    except Exception as e:
        return back_to_queue(f"Review not saved: database error ({e}).", False)

    return back_to_queue(f"Review saved for #{ticket_id}: {result}.", True)


@app.route("/queue/api/acknowledge", methods=["POST"])
def acknowledge_update_api():
    """Acknowledge the current cached update without contacting Freshdesk."""
    config = filters_from_args(request.form)
    def back(msg, ok):
        session["flash"] = ("ok" if ok else "err", msg)
        return redirect(f"/queue?{filter_query_string(config)}", code=303)
    if not csrf_valid(request.form.get("csrf_token")):
        return back("Update not acknowledged: invalid security token. Reload the page and try again.", False)
    try:
        ticket_id = int(request.form.get("ticket_id", ""))
    except (TypeError, ValueError):
        return back("Update not acknowledged: invalid ticket ID.", False)
    try:
        pool, _ = get_ticket_pool()
        ticket = next((t for t in pool if t.get("id") == ticket_id), None)
        state = load_review_rows().get(ticket_id)
        if ticket is None or state is None or state.get("review_result") not in REVIEWED_STATES:
            return back("Update not acknowledged: reviewed ticket is unavailable.", False)
        current = ticket.get("updated_at")
        if parse_dt(current) is None:
            return back("Update not acknowledged: cached timestamp is invalid.", False)
        if not updated_since_review(ticket, state):
            return back(f"Update already acknowledged for #{ticket_id}.", True)
        if not _advance_review_snapshot(ticket_id, current):
            return back("Update not acknowledged: local review state was unavailable.", False)
    except Exception:
        return back("Update not acknowledged: local state could not be updated.", False)
    return back(f"Update acknowledged for #{ticket_id}.", True)


@app.route("/queue/api/opened", methods=["POST"])
def opened_api():
    """Record that a ticket link was opened (JSON). The click handler calls
    this without preventing the new tab, so the Freshdesk ticket always opens
    even if recording fails. Success/failure is reported honestly: the UI only
    claims the mark was saved when this endpoint returns ok."""
    token = request.headers.get("X-CSRF-Token") or (request.get_json(silent=True) or {}).get("csrf_token")
    if not csrf_valid(token):
        return jsonify({"ok": False, "error": "invalid security token"}), 403

    body = request.get_json(silent=True) or {}
    raw_ticket_id = body.get("ticket_id")
    if raw_ticket_id is None:
        return jsonify({"ok": False, "error": "missing ticket_id"}), 400
    try:
        ticket_id = int(raw_ticket_id)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "invalid ticket_id"}), 400

    try:
        pool, _ = get_ticket_pool()
        known = any(t.get("id") == ticket_id for t in pool)
    except Exception:
        known = False
    if not known:
        return jsonify({"ok": False, "error": "unknown ticket id"}), 404

    try:
        result = mark_opened(ticket_id)
    except Exception as e:
        return jsonify({"ok": False, "error": f"database error: {e}"}), 500

    return jsonify({
        "ok": True,
        "review_result": result,
        "last_opened_id": last_opened_ticket_id(),
    })


def resolve_bind_host(host):
    """Refuse unsafe external binds. The scanner must stay on the loopback
    interface. Raises SystemExit for 0.0.0.0 so misuse is loud and immediate.
    """
    if host == "0.0.0.0":
        raise SystemExit(
            "Refusing to bind to 0.0.0.0. Set HOST=127.0.0.1 or export PORT=5050."
        )
    return host


# ---------------------------------------------------------------------------
# Dashboard template
# ---------------------------------------------------------------------------

# Shared application theme (design source of truth = the /queue theme).
# Both QUEUE_HTML and CLOSED_HTML reference this single stylesheet so the two
# pages share body background, content width, typography, navigation, panels,
# buttons, form controls, tables, badges, links, focus-visible and responsive
# breakpoints. Navigation is rendered by _nav_html(current).
_SHARED_CSS = """
 :root{--fd-customer-responded:#09218D;--fd-customer-responded-text:#FFFFFF;--fd-waiting-customer:#E9AE3D;--fd-waiting-customer-text:#1A1A1A;--fd-last-opened:#6A1B9A;--fd-last-opened-text:#FFFFFF}
 body{font-family:system-ui,Arial,sans-serif;max-width:1440px;margin:auto;padding:16px;background:#f5f5f5;color:#222}
  h1{font-size:22px;margin:0 0 4px}
  .sub{color:#666;font-size:13px;margin-bottom:16px}
  .queue-status{display:inline-flex;gap:7px;margin:0 0 14px}
  .status-chip{border-radius:999px;padding:4px 9px;font-size:12px;font-weight:650;background:#e8f5e9;color:#245b31}
  .status-chip.readonly{background:#eef1f5;color:#4d5968}
  .workflow-tabs{display:flex;gap:4px;flex-wrap:wrap;border-bottom:1px solid #ddd;margin:18px 0 14px}
  .workflow-tab{padding:9px 13px;border:1px solid transparent;border-bottom:3px solid transparent;color:#4b5563;text-decoration:none;font-weight:650;font-size:13px}
  .workflow-tab:hover,.workflow-tab.active{color:#173b72;background:#fff;border-color:#e1e5ea;border-bottom-color:#2f6fca}
  .workflow-tab-count{font-weight:500;color:#6b7280}
  .updated-section{margin-top:10px}
  .workflow-section-label{font-size:12px;letter-spacing:.08em;font-weight:750;color:#536273;margin:16px 0 7px}
  .live-meta{font-size:13px;color:#667085;margin:0 0 12px}
  .coverage-warning{margin:0 0 12px;padding:8px 11px;border:1px solid #e0c060;border-radius:6px;background:#fff8df;font-size:13px;color:#624d12}
  .empty{padding:24px;text-align:center;color:#667085;background:#fff;border:1px dashed #d5dbe3;border-radius:8px}
  /* Keep the range editor's native hidden state authoritative.  The
     !important guard prevents a later layout rule from masking it. */
  [hidden]{display:none!important}
  .custom-days.hidden{display:none}
 .banner{background:#fff3cd;border:1px solid #e0c060;padding:8px 12px;border-radius:6px;font-size:13px;margin-bottom:14px}
 .banner.err{background:#fdecea;border-color:#d66;color:#8a1f1f}
 .banner.ok{background:#e8f5e9;border-color:#6a9;color:#1e4d2b}
 .controls{background:#fff;border:1px solid #e0e0e0;border-radius:10px;padding:16px 18px;margin-bottom:10px;box-shadow:0 1px 2px rgba(0,0,0,.04)}
 .queue-controls{display:grid;grid-template-columns:minmax(270px,38fr) minmax(0,62fr);gap:20px;padding:16px 18px}
 .queue-data-area{min-width:0;padding-right:20px;border-right:1px solid #e7e9ed}
 .queue-filter-controls{min-width:0}
 .queue-controls .controls{background:transparent;border:0;border-radius:0;box-shadow:none;padding:0;margin:0}
 .queue-controls .panel-region{padding-bottom:12px;margin-bottom:12px}
 .queue-data-area .panel-region{display:flex;flex-wrap:wrap;align-items:center;gap:10px;padding-bottom:0;margin-bottom:8px;border-bottom:0}
 .queue-card-heading{font-size:11px;font-weight:750;letter-spacing:.08em;color:#536273;margin:0 0 9px}
 .queue-data-area .live-meta{margin:0 0 8px}
 .reconcile-details{margin-top:8px;border-top:1px solid #f0f0f0;padding-top:8px}
 .reconcile-details summary{cursor:pointer;color:#315f9d;font-size:13px;font-weight:650;list-style-position:inside}
 .reconcile-details[open] summary{margin-bottom:10px}
 .reconcile-details .reconcile-panel{padding:0;border-bottom:0;margin-bottom:0}
 .refresh-status{display:none;font-size:13px;line-height:1.4;margin:8px 0 0;padding:7px 9px;border-radius:6px;background:#f1f3f4;color:#44505f}
 .refresh-status:not(:empty){display:block}
 .refresh-status.success{background:#e8f5e9;color:#1e4d2b}
 .refresh-status.warning{background:#fff8df;color:#624d12}
 .refresh-status.error{background:#fdecea;color:#8a1f1f}
 .queue-cancel{padding:7px 15px;font-size:13px;font-weight:500;color:#5f6368;background:#fff;border:1px solid #c6c9cf;border-radius:6px;cursor:pointer}
 .queue-card-footer{grid-column:1 / -1;border-top:1px solid #edf0f3;padding-top:10px;color:#7a8491;font-size:12px}
 .controls .panel-region{display:flex;flex-wrap:wrap;align-items:center;gap:16px;padding-bottom:14px;margin-bottom:14px;border-bottom:1px solid #f0f0f0}
 .controls .panel-region:last-child{padding-bottom:0;margin-bottom:0;border-bottom:0}
 .queue-controls .queue-data-area .panel-region{padding-bottom:0;margin-bottom:8px;border-bottom:0}
 .queue-controls .queue-filter-controls .panel-region{padding-bottom:12px;margin-bottom:12px}
 .region-time{justify-content:space-between}
 .days-field{display:inline-flex;align-items:center;gap:7px;flex-wrap:wrap}
 .days-field .lbl{font-size:13px;color:#444;white-space:nowrap}
 .controls input[type=number]{width:76px;padding:6px 9px;font-size:14px;border:1px solid #bdbdbd;border-radius:6px;background:#fff}
 .controls input[type=number]:focus-visible{outline:2px solid #1a73e8;outline-offset:1px}
 .preset-group{display:inline-flex;flex-wrap:wrap;gap:4px;align-items:center;background:#f3f4f6;border:1px solid #e5e7eb;border-radius:8px;padding:4px}
 .preset{font-size:12px;font-weight:600;color:#444;text-decoration:none;padding:5px 11px;border-radius:6px;line-height:1;letter-spacing:.02em}
 .preset:hover{background:#e5e7eb;color:#111}
 .preset[aria-current=page]{background:#1a73e8;color:#fff;box-shadow:inset 0 0 0 1.5px #175cd3}
 .preset[aria-current=page] .preset-mark{font-weight:700}
 .preset:focus-visible{outline:2px solid #1a73e8;outline-offset:2px}
 .region-groups{display:flex;flex-wrap:wrap;gap:12px}
 .filter-group{border:1px solid #e4e6eb;border-radius:8px;padding:10px 12px 11px;margin:0;min-width:150px;flex:1 1 200px;display:flex;flex-direction:column;gap:7px;background:#fcfcfd}
 .filter-group .group-lbl{font-size:11px;font-weight:700;color:#5f6368;text-transform:uppercase;letter-spacing:.5px;padding:0 2px}
 .field{display:inline-flex;align-items:center;gap:5px}
 .field .lbl{font-size:13px;color:#444;white-space:nowrap}
 .filter-group .field label{display:inline-flex;align-items:center;gap:6px;cursor:pointer;font-size:13px;color:#222;text-transform:none;letter-spacing:0;font-weight:400}
 .filter-group .field input[type=checkbox]{width:16px;height:16px;accent-color:#1a73e8;cursor:pointer}
 .filter-group .field input[type=checkbox]:focus-visible{outline:2px solid #1a73e8;outline-offset:2px}
 .filter-group .field-hint{font-size:11px;color:#8a8f98;line-height:1.4;margin:0}
 .region-actions{display:flex;flex-wrap:wrap;align-items:center;gap:14px;justify-content:space-between}
 .view-field{display:inline-flex;align-items:center;gap:8px}
 .view-field label{font-size:13px;color:#444;white-space:nowrap}
 .controls select{padding:7px 10px;font-size:13px;border:1px solid #bdbdbd;border-radius:6px;background:#fff}
 .controls select:focus-visible{outline:2px solid #1a73e8;outline-offset:2px}
 .action-buttons{display:inline-flex;gap:10px;flex-wrap:wrap;align-items:center}
 .controls button[type=submit]{padding:8px 18px;font-size:13px;font-weight:600;border:1px solid #1565c0;background:#1a73e8;color:#fff;border-radius:6px;cursor:pointer}
 .controls button[type=submit]:hover{background:#1664d0}
 .controls button[type=submit]:focus-visible{outline:2px solid #0d47a1;outline-offset:2px}
 .controls a.reset{display:inline-block;padding:7px 15px;font-size:13px;font-weight:500;color:#5f6368;background:#fff;border:1px solid #c6c9cf;border-radius:6px;text-decoration:none}
 .controls a.reset:hover{border-color:#9aa0a6;color:#202124}
 .controls a.reset:focus-visible{outline:2px solid #1a73e8;outline-offset:2px}
 .filter-summary{font-size:13px;color:#3c4043;background:#f1f3f4;border:1px solid #e0e0e0;border-radius:8px;padding:7px 12px;margin:8px 0 12px}
 @media (max-width:720px){.controls{padding:14px}.queue-controls{grid-template-columns:1fr;gap:16px;padding:14px}.queue-data-area{padding:0 0 16px;border-right:0;border-bottom:1px solid #e7e9ed}.region-time{flex-direction:column;align-items:flex-start;gap:10px}.region-groups{flex-direction:column}.filter-group{flex:1 1 auto;min-width:0}.region-actions{flex-direction:column;align-items:flex-start;gap:12px}.action-buttons{width:100%;justify-content:space-between}.controls button[type=submit]{flex:1 1 auto}.controls a.reset{flex:1 1 auto;text-align:center}}
 .count{font-size:13px;color:#555;margin-bottom:8px}
 .tablewrap{overflow-x:auto;background:#fff;border:1px solid #ddd;border-radius:8px}
 table{border-collapse:collapse;width:100%;font-size:13px;min-width:1180px}
 #queue-table th:nth-child(2),#queue-table td:nth-child(2){min-width:250px}
 #queue-table th:nth-child(9),#queue-table td:nth-child(9){min-width:240px}
 td.tags-cell{white-space:normal;overflow-wrap:anywhere;word-break:break-word;line-height:1.45}
 th,td{text-align:left;padding:8px 10px;border-bottom:1px solid #eee;vertical-align:top}
 th{background:#fafafa;font-size:12px;color:#666;white-space:nowrap}
 tr.rv-unreviewed{background:#fff}
 tr.rv-opened{background:#fff8e1}
 tr.rv-opened td:first-child{box-shadow:inset 4px 0 0 #f9a825}
 tr.rv-resolved{background:#e8f5e9}
 tr.rv-na{background:#eeeeee}
 tr.rv-none{background:#e3f2fd}
 tr.rv-followup{background:#fff3e0}
 tr.rv-last-opened{outline:3px solid var(--fd-last-opened);outline-offset:-3px}
 tr.rv-last-opened td:first-child{box-shadow:inset 4px 0 0 var(--fd-last-opened)}
 .b-last-opened{background:var(--fd-last-opened);color:var(--fd-last-opened-text)}
 a.tid{font-weight:bold;color:#1565c0;text-decoration:none}
 a.tid:hover{text-decoration:underline}
 a.sbj{color:#222;text-decoration:none}
 a.sbj:hover{text-decoration:underline}
 .badges{display:flex;flex-wrap:wrap;gap:4px}
 .badge{font-size:11px;font-weight:bold;padding:2px 6px;border-radius:4px;white-space:nowrap;letter-spacing:.02em}
 .b-review{border:1px solid #bbb;color:#333;background:#fff}
 .b-review.rv-opened{background:#fff8e1;border-color:#f9a825;color:#5d4037}
 .b-overdue{background:#d32f2f;color:#fff}
 .b-responded{background:var(--fd-customer-responded);color:var(--fd-customer-responded-text)}
 .b-waiting{background:var(--fd-waiting-customer);color:var(--fd-waiting-customer-text)}
  .b-missing{background:#757575;color:#fff}
  .b-triage{display:inline-block;margin:0 3px 3px 0;background:#fff3cd;border:1px solid #d39e00;color:#664d03;white-space:normal}
  .triage-reasons{min-width:175px}
 .b-sla{background:#e65100;color:#fff}
 .b-updated{background:#00838f;color:#fff}
 .b-closed{background:#00838f;color:#fff}
 .toast{position:fixed;right:16px;bottom:16px;max-width:340px;background:#fdecea;border:1px solid #d66;color:#8a1f1f;padding:10px 14px;border-radius:6px;font-size:13px;box-shadow:0 2px 8px rgba(0,0,0,.15);z-index:99}
 .toast.hidden{display:none}
 .meta{color:#666;white-space:nowrap}
 .empty{color:#666;padding:24px;text-align:center;font-size:14px}
 .rvform{margin:0}
 .rvform select{padding:4px 6px;font-size:12px;border:1px solid #bbb;border-radius:4px;max-width:150px}
 .foot{color:#999;font-size:11px;margin-top:10px}
 .top-nav{display:flex;flex-wrap:wrap;gap:6px;align-items:center;margin:0 0 18px;padding:0 0 12px;border-bottom:1px solid #d0d7de}
 .top-link{display:inline-block;padding:7px 14px;font-size:13px;font-weight:600;color:#3c4043;background:#fff;border:1px solid #c6c9cf;border-radius:999px;text-decoration:none;white-space:nowrap}
 .top-link:hover{border-color:#9aa0a6;background:#f1f3f4;color:#202124}
 .top-link[aria-current=page]{background:#1a73e8;border-color:#1565c0;color:#fff}
 .top-link:focus-visible{outline:2px solid #1a73e8;outline-offset:2px}
 @media (max-width:500px){.top-nav{gap:6px}.top-link{white-space:normal;text-align:center;flex:1 1 auto}}
"""


def _nav_html(current):
    """Shared top navigation (pill/tab style matching the queue theme).

    Renders the two dashboard links with the active page marked via
    aria-current=page. The .top-nav/.top-link rules live in _SHARED_CSS, so the
    two links are visibly spaced (flex + gap) instead of running together.
    """
    def link(href, label, active):
        attr = ' aria-current="page"' if active else ""
        return '<a class="top-link" href="%s"%s>%s</a>' % (href, attr, label)
    return (
        '<nav class="top-nav" aria-label="Dashboard pages">'
        + link("/queue", "Review Queue", current == "queue")
        + link("/closed", "Closed Ticket Housekeeping", current == "closed")
        + "</nav>"
    )


QUEUE_HTML = """\
<!doctype html><html lang=en><head><meta charset=utf-8>
<meta name=viewport content="width=device-width,initial-scale=1">
<title>Freshdesk Review Queue</title>
<style>{{ shared_css|safe }}</style></head><body>
<h1>Review Queue</h1>
<div class=queue-status><span class=status-chip>{{ '● Offline' if offline else '● Live' }}</span><span class="status-chip readonly">Read-only to Freshdesk</span></div>
{% if offline %}<div class=sub><strong>OFFLINE MODE</strong> — Offline fixture data · no network access.</div>{% elif live_cache_missing %}<div class=sub>Live mode — no Freshdesk data retrieved yet.</div>{% endif %}

{% if flash %}
<div class="banner {{ 'ok' if flash[0] == 'ok' else 'err' }}" role=status>{{ flash[1] }}</div>
{% endif %}
{% if error %}
<div class="banner err" role=alert>{{ error }}</div>
{% endif %}
<section class="controls queue-controls" aria-label="Queue controls">
<div class=queue-data-area>
<h2 class=queue-card-heading>DATA</h2>
<form class=refresh-controls method=post action=/queue/api/refresh novalidate id=queue-refresh-form>
  <input type=hidden name=csrf_token value="{{ csrf_token }}">
  <input type=hidden name=mode value="normal" id=refresh-mode>
  <div class="panel-region region-time"><div class="action-buttons"><button type=submit class=apply id=queue-refresh>Refresh Tickets</button><button type=button id=queue-cancel class=queue-cancel hidden>Cancel</button></div></div>
    <p class=live-meta>Last refreshed {{ last_refresh_display }} · {{ cached_ticket_count }} tickets cached</p>
    <p class=live-meta id=auto-refresh-status>Auto refresh: {% if not auto_refresh.enabled %}Off{% elif auto_refresh.seconds_until_next is none or auto_refresh.seconds_until_next <= 60 %}On · due soon{% else %}On · next in {{ (auto_refresh.seconds_until_next / 60)|round(0, 'ceil')|int }} min{% endif %}</p>

   {% if live_cache_missing %}<p class=field-hint>No cache baseline yet; Refresh Tickets will initialize it.</p>{% endif %}
   <details class=reconcile-details>
    <summary>Reconcile history</summary>
    <div class="panel-region region-time reconcile-panel">
      <strong>Reconcile Range</strong>
      <p class=field-hint>Re-check historical data and merge it into the existing cache; it does not replace your cache or local review history.</p>
      <div class=preset-group role=group aria-label="Retrieval range">
        {% for d in [7, 14, 30, 60, 90] %}<a {% if config.days == d %}class="preset preset-on active"{% else %}class=preset{% endif %} href="/queue?{{ preset_urls[d] }}"{% if config.days == d %} aria-current=page{% endif %}>{{ d }}d</a>{% endfor %}
        <button type=button class="preset{% if config.days not in [7,14,30,60,90] %} active{% endif %}" id=custom-days-toggle{% if config.days not in [7,14,30,60,90] %} aria-current=page{% endif %} aria-pressed="{{ 'true' if config.days not in [7,14,30,60,90] else 'false' }}">Custom…</button>
        <span class=custom-days id=custom-days-wrap{% if config.days in [7,14,30,60,90] %} hidden{% endif %}><label class=lbl for=custom-days>Days</label><input id=custom-days type=number name=days min={{ days_min }} max={{ days_max }} value="{{ config.days }}" aria-label="Custom days" step=1></span>
      </div>
      <div class=action-buttons><button type=button class=apply id=queue-reconcile>Reconcile Range</button></div>
    </div>
   </details>
</form>
<div id=queue-refresh-status class=refresh-status role=status aria-live=polite></div>
</div>
<form class="controls queue-filter-controls" method=get action=/queue novalidate id=queue-filter-form data-rendered-mode="{{ config.mode }}">
   <h2 class=queue-card-heading>REVIEW FILTERS</h2>
    <input type=hidden name=days value="{{ config.days }}">
    <input type=hidden name=queue_scope value="{{ config.get("queue_scope", "main") }}">
   {% if config.mode == 'closed' and config.normal_return is defined %}{% for name, value in config.normal_return.items() %}<input type=hidden name="normal_{{ name }}" value="{{ '1' if value is sameas true else '0' if value is sameas false else value }}">{% endfor %}{% endif %}
   <div class="panel-region region-groups">
    <fieldset class="filter-group scope-group">
      <legend class=group-lbl>Review Mode</legend>
      <div class=field><label for=queue-mode>Mode</label><select id=queue-mode name=mode><option value=normal {{ 'selected' if config.mode == 'normal' }}>Normal Review</option><option value=closed {{ 'selected' if config.mode == 'closed' }}>Closed Ticket Housekeeping</option></select></div>
      {% if config.mode == 'closed' %}
      <input type=hidden name=hide_reviewed_tags value=0>
      <p class=field-hint>Closed tickets from the current 60-day master queue cache only. No separate retrieval or closed-date window.</p>
      <div class=field><input type=hidden name=photo_video_only value=0><label for=filter-photo-video><input type=checkbox id=filter-photo-video name=photo_video_only value=1 {{ 'checked' if config.photo_video_only }}> Photo/Video Review Scope</label></div>
      <div class=field><input type=hidden name=missing_tags value=0><label for=filter-missing><input type=checkbox id=filter-missing name=missing_tags value=1 {{ 'checked' if config.missing_tags }}> Missing Tags Only</label></div>
      {% else %}
      <div class=field><input type=hidden name=photo_video_only value=0><label for=filter-photo-video><input type=checkbox id=filter-photo-video name=photo_video_only value=1 {{ 'checked' if config.photo_video_only }}> Photo/video subjects only</label></div>
      <div class=field><input type=hidden name=hide_reviewed_tags value=0><label for=filter-hide-reviewed><input type=checkbox id=filter-hide-reviewed name=hide_reviewed_tags value=1 {{ 'checked' if config.hide_reviewed_tags }}> Hide tickets with reviewed/closed tags</label></div>
      <div class=field><input type=hidden name=missing_tags value=0><label for=filter-missing><input type=checkbox id=filter-missing name=missing_tags value=1 {{ 'checked' if config.missing_tags }}> Missing Tags</label></div>
      <p class=field-hint>Default working review queue. Use the mode selector for Closed Ticket Housekeeping.</p>
      {% endif %}
    </fieldset>
  </div>
  <div class="panel-region region-actions">
    <input type=hidden name=workflow_tab value="{{ config.workflow_tab }}">
    <div class=action-buttons>
       <button type=submit class=apply>Apply Filters</button>
       <a class=reset href="/queue/export.xlsx?{{ filter_query_string(config) }}" role=button aria-label="Export the current queue view">Export Current View</a>
       <a class=reset href="/queue?mode=normal&amp;photo_video_only=1&amp;hide_reviewed_tags=1&amp;overdue=0&amp;responded=0&amp;waiting=0&amp;missing_tags=0&amp;days={{ config.days }}&amp;review_view=all&amp;workflow_tab=main" role=button aria-label="Reset to the default Review Scope">Reset to Default Review Scope</a>

      <a class=reset href="/queue?mode=normal&amp;photo_video_only=0&amp;hide_reviewed_tags=0&amp;overdue=0&amp;responded=0&amp;waiting=0&amp;missing_tags=0&amp;days={{ config.days }}&amp;review_view=all&amp;workflow_tab=main" role=button aria-label="Show every cached ticket">Show All Cached Tickets</a>
    </div>
  </div>
</form>
<div class=queue-card-footer>Manual refresh only · Local filters never contact Freshdesk.</div>
</section>
<p class=filter-summary role=status>{{ active_summary }}</p>

{% if config.mode == 'normal' %}
<nav class=workflow-tabs aria-label="Queue scope">
{% for scope, label in [('main', 'Main Queue'), ('triage', 'Needs Triage')] %}<a class="workflow-tab{% if config.get("queue_scope", "main") == scope %} active{% endif %}" href="/queue?{{ filter_query_string(dict(config, queue_scope=scope)) }}" {% if config.get("queue_scope", "main") == scope %}aria-current=page{% endif %}>{{ label }} <span class=workflow-tab-count>({{ queue_scope_counts[scope] }})</span></a>{% endfor %}
</nav>
{% endif %}
<nav class=workflow-tabs aria-label="Review workflow">
{% for tab in ['main','supervisor','followup','resolved','no_action'] %}<a class="workflow-tab{% if config.workflow_tab == tab %} active{% endif %}" href="/queue?{{ filter_query_string(dict(config, workflow_tab=tab)) }}" {% if config.workflow_tab == tab %}aria-current=page{% endif %}>{{ workflow_labels[tab] }} <span class=workflow-tab-count>({{ workflow_counts[tab] }})</span></a>{% endfor %}
</nav>
<p class=count>{{ total }} tickets displayed from the current cache</p>
{% if last_opened_id is not none %}
  {% if last_opened_rendered %}
<p class=last-opened-bar><button type=button id=last-opened-jump aria-controls=queue-table>Jump to Last Opened</button></p>
  {% else %}
<div class="banner" id=last-opened-hidden role=status>Last opened ticket is hidden by the current filters.</div>
  {% endif %}
{% endif %}
{% if tickets %}
{% set updated_tickets = tickets|selectattr('updated_flag')|list %}
{% set ordinary_tickets = tickets|rejectattr('updated_flag')|list %}
{% if config.workflow_tab == 'main' and updated_tickets %}<div class=workflow-section-label>UPDATED SINCE REVIEW</div>{% endif %}
<div class=tablewrap>
<table id=queue-table>
<caption class=visually-hidden>Freshdesk review queue</caption>
<tr>
  <th scope=col>Ticket</th><th scope=col>Subject</th><th scope=col>Status</th>
   <th scope=col>Badges</th>{% if config.mode == 'normal' and config.get("queue_scope", "main") == 'triage' %}<th scope=col>Why here?</th>{% endif %}<th scope=col>Review</th><th scope=col>Due / SLA</th>
  <th scope=col>Updated</th><th scope=col>Created</th><th scope=col>Tags</th>
</tr>
{% for t in tickets %}
<tr class="{{ t.row_class }}" data-ticket-id="{{ t.id }}">
  <td><a class="tid fd-link" href="{{ t.url }}" target=_blank rel="noopener noreferrer" data-ticket-id="{{ t.id }}" aria-label="Open ticket #{{ t.id }} in Freshdesk (new tab)">#{{ t.id }}</a></td>
  <td><a class="sbj fd-link" href="{{ t.url }}" target=_blank rel="noopener noreferrer" data-ticket-id="{{ t.id }}" aria-label="Open subject of ticket #{{ t.id }} in Freshdesk (new tab)">{{ t.subject }}</a></td>
  <td>{{ t.status_label }}</td>
   <td><div class=badges>{% if t.last_opened %}<span class="badge b-last-opened">LAST OPENED</span>{% endif %}{% for kind, text, cls in t.badges %}<span class="badge {{ cls }}">{{ text }}</span>{% endfor %}</div></td>
   {% if config.mode == 'normal' and config.get("queue_scope", "main") == 'triage' %}<td class=triage-reasons>{% for reason in t.triage_reasons %}<span class="badge b-triage">{{ reason }}</span>{% endfor %}</td>{% endif %}
   <td>
    <form class=rvform method=post action=/queue/api/review>
      <input type=hidden name=csrf_token value="{{ csrf_token }}">
      <input type=hidden name=ticket_id value="{{ t.id }}">
      <input type=hidden name=mode value="{{ config.mode }}">
      {% if config.mode == 'closed' and config.normal_return is defined %}{% for name, value in config.normal_return.items() %}<input type=hidden name="normal_{{ name }}" value="{{ '1' if value is sameas true else '0' if value is sameas false else value }}">{% endfor %}{% endif %}
      <input type=hidden name=photo_video_only value="{{ '1' if config.photo_video_only else '0' }}">
      <input type=hidden name=hide_reviewed_tags value="{{ '1' if config.hide_reviewed_tags else '0' }}">
      <input type=hidden name=overdue value="{{ '1' if config.overdue else '0' }}">
      <input type=hidden name=responded value="{{ '1' if config.responded else '0' }}">
      <input type=hidden name=waiting value="{{ '1' if config.waiting else '0' }}">
      <input type=hidden name=missing_tags value="{{ '1' if config.missing_tags else '0' }}">
      <input type=hidden name=days value="{{ config.days }}">
       <input type=hidden name=review_view value="{{ config.review_view }}">
        <input type=hidden name=workflow_tab value="{{ config.workflow_tab }}">
        <input type=hidden name=queue_scope value="{{ config.get("queue_scope", "main") }}">
       <select name=review_result aria-label="Review result for ticket {{ t.id }}" onchange="this.form.submit()">
        {% for s in review_states %}<option value="{{ s }}" {{ 'selected' if t.result == s }}>{{ s }}</option>{% endfor %}
       </select>
       {% if t.can_acknowledge %}<button type=submit class=acknowledge formaction=/queue/api/acknowledge formmethod=post aria-label="Acknowledge Update for ticket {{ t.id }}">Acknowledge Update</button>{% endif %}
     </form>
   </td>
  <td class=meta>{{ t.due_display | safe }}</td>
  <td class=meta>{{ t.updated_display }}</td>
  <td class=meta>{{ t.created_display }}</td>
  <td class=tags-cell>{% if t.tags %}{{ t.tags|join(', ') }}{% else %}<em style=color:#bbb>none</em>{% endif %}</td>
</tr>
{% endfor %}
</table>
</div>
{% else %}
<div class=empty>{% if config.workflow_tab == 'main' %}No tickets need review.{% elif config.workflow_tab == 'supervisor' %}No tickets are waiting for supervisor review.{% elif config.workflow_tab == 'followup' %}No tickets need follow-up.{% elif config.workflow_tab == 'resolved' %}No resolved tickets in current view.{% else %}No no-action tickets in current view.{% endif %}</div>
{% endif %}

<div class=foot>Review results are stored locally only (SQLite) and are never sent to Freshdesk. Ticket links open Freshdesk in a new tab; opening a ticket marks it as Opened / In Review locally.</div>

<script>
var CSRF_TOKEN = {{ csrf_token_json | safe }};
var REVIEW_CLASS = {
  'Unreviewed': 'rv-unreviewed',
  'Opened / In Review': 'rv-opened',
  'Resolved': 'rv-resolved',
  'Not Applicable to Me': 'rv-na',
  'No Action Needed': 'rv-none',
  'Needs Follow-Up': 'rv-followup'
};
function badgeText(result) { return result === 'Opened / In Review' ? 'OPENED / IN REVIEW' : result; }
var toastEl = null;
function showError(msg) {
  if (!toastEl) {
    toastEl = document.createElement('div');
    toastEl.className = 'toast hidden';
    toastEl.setAttribute('role', 'status');
    document.body.appendChild(toastEl);
  }
  toastEl.textContent = msg;
  toastEl.classList.remove('hidden');
  clearTimeout(showError._t);
  showError._t = setTimeout(function () { toastEl.classList.add('hidden'); }, 6000);
}
// The Last Opened focus marker is derived server-side from the newest valid
// last_opened_at (spec section 3/6) and reported back by the opened endpoint
// as last_opened_id. On a *confirmed* save we move the marker in the DOM
// without reloading: strip the marker class/badge from every row (idempotent,
// so repeat clicks never duplicate the badge), then apply it to the target row
// (if rendered) and keep the jump control / hidden-message in sync.
function moveLastOpened(newId) {
  if (typeof newId === 'undefined' || newId === null) { return; }
  ensureLastOpenedChrome();
  document.querySelectorAll('.b-last-opened').forEach(function (b) {
    b.parentNode.removeChild(b);
  });
  document.querySelectorAll('tr.rv-last-opened').forEach(function (r) {
    r.classList.remove('rv-last-opened');
  });
  var target = document.querySelector('tr[data-ticket-id="' + newId + '"]');
  var bar = document.querySelector('.last-opened-bar');
  var hidden = document.getElementById('last-opened-hidden');
  if (target) {
    target.classList.add('rv-last-opened');
    var badges = target.querySelector('.badges');
    if (badges) {
      var span = document.createElement('span');
      span.className = 'badge b-last-opened';
      span.textContent = 'LAST OPENED';
      badges.appendChild(span);
    }
    if (bar) { bar.style.display = ''; }
    if (hidden) { hidden.style.display = 'none'; }
  } else {
    if (bar) { bar.style.display = 'none'; }
    if (hidden) { hidden.style.display = ''; }
  }
}
// Jump to Last Opened: smooth-scroll to the marked row and set temporary
// keyboard focus. Never opens the Freshdesk link and never makes a network
// request (spec section 5).
function jumpToLastOpened() {
  var row = document.querySelector('tr.rv-last-opened');
  if (!row) { return; }
  row.scrollIntoView({behavior: 'smooth', block: 'center'});
  row.setAttribute('tabindex', '-1');
  row.focus({preventScroll: true});
  setTimeout(function () { row.removeAttribute('tabindex'); }, 1500);
}
// The jump bar and the hidden-by-filters notice are server-rendered only when
// a last-opened marker already existed at page load. When the *first* click
// creates the marker client-side, build the same chrome here so the jump
// control appears without a reload.
function ensureLastOpenedChrome() {
  if (!document.querySelector('.last-opened-bar')) {
    var bar = document.createElement('p');
    bar.className = 'last-opened-bar';
    var btn = document.createElement('button');
    btn.type = 'button';
    btn.id = 'last-opened-jump';
    btn.setAttribute('aria-controls', 'queue-table');
    btn.textContent = 'Jump to Last Opened';
    bar.appendChild(btn);
    var anchor = document.querySelector('.tablewrap') || document.body;
    anchor.parentNode.insertBefore(bar, anchor);
  }
  if (!document.getElementById('last-opened-hidden')) {
    var div = document.createElement('div');
    div.className = 'banner';
    div.id = 'last-opened-hidden';
    div.setAttribute('role', 'status');
    div.textContent = 'Last opened ticket is hidden by the current filters.';
    var divAnchor = document.querySelector('.tablewrap') || document.body;
    divAnchor.parentNode.insertBefore(div, divAnchor);
  }
}
// Delegated listener: works whether the button was server-rendered at load or
// created later by ensureLastOpenedChrome() after the first marker click.
document.addEventListener('click', function (e) {
  if (e.target && e.target.id === 'last-opened-jump') { jumpToLastOpened(); }
});
// both the ticket-number and subject links. The default navigation is never
// prevented: the Freshdesk ticket always opens in a new tab synchronously,
// and the local state request runs in the background (spec section 5).
document.querySelectorAll('a[data-ticket-id]').forEach(function (a) {
  a.addEventListener('click', function () {
    var tid = this.getAttribute('data-ticket-id');
    if (!tid) { return; }
    var row = this.closest('tr');
    fetch('/queue/api/opened', {
      method: 'POST',
      headers: {'Content-Type': 'application/json', 'X-CSRF-Token': CSRF_TOKEN},
      body: JSON.stringify({ticket_id: tid})
    }).then(function (r) { return r.json(); }).then(function (d) {
      if (d && d.ok && d.review_result) {
        // Confirmed saved: highlight the row and update badge + selector.
        var cls = REVIEW_CLASS[d.review_result] || 'rv-unreviewed';
        if (row) {
          var extra = [];
          row.classList.forEach(function (c) { if (c.indexOf('rv-') !== 0) { extra.push(c); } });
          row.className = extra.concat([cls]).join(' ');
        }
        var badge = row ? row.querySelector('.b-review') : null;
        if (badge) { badge.className = 'badge b-review ' + cls; badge.textContent = badgeText(d.review_result); }
        var sel = row ? row.querySelector('select[name=review_result]') : null;
        if (sel && sel.querySelector('option[value="' + d.review_result + '"]')) { sel.value = d.review_result; }
        // Marker only moves on a confirmed save (spec: save failure must not
        // falsely move the Last Opened marker).
        moveLastOpened(d.last_opened_id);
      } else {
        showError('Could not save Opened / In Review state for #' + tid + ' (not saved).');
      }
    }).catch(function () {
      showError('Could not save Opened / In Review state for #' + tid + ' (not saved).');
    });
  });
});
// Local Apply Filters form (GET /queue): canonicalize every checkbox to an
// explicit 0/1 so every local checkbox can remain intentionally unchecked. This navigation is GET-only and never contacts Freshdesk.
(function () {
  var form = document.getElementById('queue-filter-form');
  if (!form) { return; }
  function normDays(raw) {
    var v = String(raw == null ? '' : raw).trim();
    if (!/^[0-9]+$/.test(v)) { return 60; }
    var n = parseInt(v, 10);
    return (n >= 1 && n <= 365) ? n : 60;
  }
  function normView(v) { return (v === 'active' || v === 'completed') ? v : 'all'; }
  // Hidden 0 values paired with the checkboxes make this a native, bookmarkable
  // GET form: checked controls submit the later 1 value, unchecked controls
  // submit only 0. No retrieval route is involved.
  // The mode selector represents two independent, URL-backed local workspaces.
  // A native submission after changing it would submit the outgoing mode's
  // hidden checkbox values (notably normal-mode missing_tags=0), accidentally
  // overriding Closed Housekeeping defaults.  Intercept only an actual mode
  // transition: save normal state under private return keys on entry to Closed,
  // and reconstruct the normal URL when returning.  Ordinary Apply submissions
  // remain native GET requests.
  var renderedMode = form.getAttribute('data-rendered-mode') || 'normal';
  var modeEl = form.querySelector('select[name="mode"]');
  var normalKeys = ['photo_video_only', 'hide_reviewed_tags', 'missing_tags', 'days', 'review_view', 'workflow_tab'];
  function currentValue(name, fallback) {
    var checked = form.querySelector('input[type=checkbox][name="' + name + '"]');
    if (checked) { return checked.checked ? '1' : '0'; }
    var input = form.querySelector('[name="' + name + '"]');
    return input ? input.value : fallback;
  }
  function normalReturnValue(q, name) {
    var canonical = {
      photo_video_only: '1',
      hide_reviewed_tags: '1',
      missing_tags: '0',
      days: '60',
      review_view: 'all',
      workflow_tab: 'main'
    };
    return q.has('normal_' + name) ? q.get('normal_' + name) : canonical[name];
  }
  form.addEventListener('submit', function (e) {
    var targetMode = modeEl ? modeEl.value : renderedMode;
    if (targetMode === renderedMode) { return; }
    e.preventDefault();
    var q = new URLSearchParams(window.location.search);
    q.set('mode', targetMode);
    if (renderedMode === 'normal' && targetMode === 'closed') {
      normalKeys.forEach(function (name) { q.set('normal_' + name, currentValue(name, name === 'workflow_tab' ? 'main' : name === 'review_view' ? 'all' : '0')); });
      // Absent is significant on Closed entry: the server supplies its ON
      // defaults, while later explicit Closed form choices remain 0/1.
      ['photo_video_only', 'hide_reviewed_tags', 'missing_tags'].forEach(function (name) { q.delete(name); });
    } else if (renderedMode === 'closed' && targetMode === 'normal') {
      normalKeys.forEach(function (name) { q.set(name, normalReturnValue(q, name)); q.delete('normal_' + name); });
    }
    q.delete('page');
    window.location.assign('/queue?' + q.toString());
  });
  function syncControlsFromURL() {
    var q = new URLSearchParams(window.location.search);
    ['photo_video_only', 'hide_reviewed_tags', 'overdue', 'responded', 'waiting', 'missing_tags'].forEach(function (n) {
      var el = form.querySelector('input[type=checkbox][name="' + n + '"]');
      // An absent URL key means the server-rendered default is authoritative.
      // Only explicit 0/1 values should override the control state.
      if (el && q.has(n)) { el.checked = (q.get(n) === '1'); }
    });
    var daysEl = form.querySelector('input[name=days]');
    if (daysEl) { daysEl.value = String(normDays(q.get('days'))); }
    var viewEl = form.querySelector('select[name=review_view]');
    if (viewEl) { viewEl.value = normView(q.get('review_view')); }
  }
  window.addEventListener('pageshow', function () { syncControlsFromURL(); });
})();
// Local Apply Filters (GET): prevent the JS refresh handler from ever calling
// it, and rely on native form GET submission. Any interception is removed here.


// Refresh Tickets form (POST /queue/api/refresh): start one background refresh
// and poll the local status endpoint once per second until it terminates.
(function () {
  var form = document.getElementById('queue-refresh-form');
  var statusEl = document.getElementById('queue-refresh-status');
  if (!form || !statusEl) { return; }
  var csrf = form.querySelector('input[name=csrf_token]');
  var refreshBtn = document.getElementById('queue-refresh');
   var cancelBtn = document.getElementById('queue-cancel');
   var autoStatusEl = document.getElementById('auto-refresh-status');
   var pollTimer = null;
   function renderAutoRefresh(autoRefresh) {
     if (!autoStatusEl) { return; }
     if (!autoRefresh || !autoRefresh.enabled) { autoStatusEl.textContent = 'Auto refresh: Off'; return; }
     var seconds = autoRefresh.seconds_until_next;
     if (seconds == null || seconds <= 60) { autoStatusEl.textContent = 'Auto refresh: On · due soon'; return; }
     autoStatusEl.textContent = 'Auto refresh: On · next in ' + Math.ceil(seconds / 60) + ' min';
   }

  function encode(obj) {
    return Object.keys(obj).map(function (k) { return encodeURIComponent(k) + '=' + encodeURIComponent(obj[k]); }).join('&');
  }
   function render(s) {
     renderAutoRefresh(s.auto_refresh);
     var p = s.progress || {};
     var running = (s.state === 'running');
     if (refreshBtn) { refreshBtn.disabled = running; }

    if (cancelBtn) { cancelBtn.hidden = !running; }
    var msg = s.message || '';
    if (running || s.state === 'running') {
      msg += ' Pages: ' + (p.pages_completed || 0);
      msg += ' · Tickets received: ' + (p.tickets_received || 0);
      msg += ' · Requests: ' + (p.request_count || 0);
      var rem = p.rate_limit_remaining;
      msg += ' · Rate limit remaining: ' + (rem == null ? 'Unknown' : rem);
      msg += ' · Elapsed: ' + (p.elapsed_seconds != null ? p.elapsed_seconds + 's' : '0s');
      if (p.wait_seconds) { msg += ' · Waiting before next request…'; }
    }
    statusEl.textContent = msg;
    statusEl.className = 'refresh-status';
    if (s.state === 'succeeded') { statusEl.classList.add('success'); }
    else if (s.state === 'failed') { statusEl.classList.add('error'); }
    else if (s.state === 'warning') { statusEl.classList.add('warning'); }
    if (running) {
      pollTimer = window.setTimeout(poll, 1000);
    } else if (s.state === 'succeeded') {
      // A data refresh must never silently carry local filters forward. Show
      // the newly completed cache in the default Main Queue while preserving
      // only the retrieval Days selection. The explicit workflow_tab=main is
      // required: workflow routing is authoritative, so without it the legacy
      // review_view=all fallback would override the Main Queue routing and
      // display reviewed tickets that belong on other workflow tabs.
      var doneDays = (s.days != null ? s.days : (form.querySelector('input[name=days]') || {}).value) || 60;
      window.location.assign('/queue?overdue=0&responded=0&waiting=0&missing_tags=0&days=' + encodeURIComponent(doneDays) + '&review_view=all&workflow_tab=main');
    } else if (s.state === 'failed' || s.state === 'cancelled') {
      // Leave the page as-is; the prior cache is untouched.
    }
  }
  function poll() {
    fetch('/queue/api/refresh/status').then(function (r) { return r.json(); }).then(render).catch(function () {
      statusEl.className = 'refresh-status error';
      statusEl.textContent = 'Unable to read refresh status.';
    });
  }
  if (cancelBtn) {
    cancelBtn.addEventListener('click', function () {
      fetch('/queue/api/refresh/cancel', { method: 'POST', headers: { 'Content-Type': 'application/x-www-form-urlencoded' }, body: encode({ csrf_token: csrf ? csrf.value : '' }) }).then(function () { poll(); });
    });
  }
   function startRefresh(mode) {
      var daysEl = form.querySelector('input[name=days]');
      var customEl = document.getElementById('custom-days');
      var body = encode({ csrf_token: csrf ? csrf.value : '', mode: mode, days: mode === 'reconcile' ? ((customEl || daysEl) ? (customEl || daysEl).value : '60') : '60' });
     statusEl.className = 'refresh-status';
     statusEl.textContent = 'Starting refresh…';
     if (refreshBtn) { refreshBtn.disabled = true; }
     fetch('/queue/api/refresh', { method: 'POST', headers: { 'Content-Type': 'application/x-www-form-urlencoded' }, body: body })
       .then(function (r) { return r.json(); }).then(function (s) { render(s); if (s.state === 'running') { poll(); } })
       .catch(function () { statusEl.className = 'refresh-status error'; statusEl.textContent = 'Refresh could not be started.'; if (refreshBtn) { refreshBtn.disabled = false; } });
   }
   form.addEventListener('submit', function (e) {
      e.preventDefault();
      startRefresh('normal');
      });
    var reconcileBtn = document.getElementById('queue-reconcile');
    if (reconcileBtn) reconcileBtn.addEventListener('click', function () { startRefresh('reconcile'); });
    var customToggle = document.getElementById('custom-days-toggle');
   var customWrap = document.getElementById('custom-days-wrap');
   var customInput = document.getElementById('custom-days');
    function selectCustomRange() {
      document.querySelectorAll('.reconcile-panel .preset').forEach(function (control) {
        control.classList.remove('active', 'preset-on');
        control.removeAttribute('aria-current');
        if (control !== customToggle) { control.setAttribute('aria-pressed', 'false'); }
      });
      customWrap.hidden = false;
      customWrap.classList.remove('hidden');
      customToggle.classList.add('active');
      customToggle.setAttribute('aria-current', 'page');
      customToggle.setAttribute('aria-pressed', 'true');
      customInput.focus();
    }
    if (customToggle && customWrap && customInput) customToggle.addEventListener('click', selectCustomRange);
})();
</script>
</body></html>
"""


CLOSED_HTML = """\
<!doctype html><html lang=en><head><meta charset=utf-8>
<meta name=viewport content="width=device-width,initial-scale=1">
<title>Closed Ticket Housekeeping</title>
<style>{{ shared_css|safe }}
 .complete{color:#176b35;font-weight:700}
 .incomplete{color:#8a1f1f;background:#fdecea;border:1px solid #d66;border-radius:6px;padding:8px 12px;font-size:13px;margin:8px 0}
 .closed-empty{color:#666;padding:24px 0;text-align:center;font-size:14px}
 .review-note{color:#666;font-size:12px;margin-top:10px}
</style></head><body>
{{ nav|safe }}
<h1>Closed Ticket Housekeeping</h1>
<p class=sub>Find closed tickets that may need housekeeping, such as tickets with no tags.</p>
{% if flash %}<div class="banner {{ 'ok' if flash[0] == 'ok' else 'err' }}" role=status>{{ flash[1] }}</div>{% endif %}
<div class="banner" role=status>{% if offline %}<strong>OFFLINE MODE — Synthetic fixture data only</strong>{% else %}<strong>LIVE CACHE MODE — Opening or reloading this page makes no Freshdesk request.</strong>{% endif %}</div>
<div class="refresh-panel" data-offline="{{ 1 if offline else 0 }}">
  {% if offline %}
  <button type=button disabled>Refresh from Freshdesk</button>
  <span class=field-hint>Live refresh is disabled in offline mode.</span>
  {% else %}
  <button type=button id=closed-refresh>Refresh from Freshdesk</button>
  <button type=button id=closed-cancel hidden>Cancel</button>
  <span id=closed-refresh-status role=status aria-live=polite>{% if live_cache %}Cached through: {{ live_cache.fetched_at }} · Coverage starts: {{ live_cache.coverage_start }}{% else %}No live Closed cache yet.{% endif %}</span>
  {% endif %}
</div>
<form class="controls" method=get action=/closed novalidate>
  <div class="panel-region region-time">
    <span class="days-field field"><span class=lbl>Closed in the last</span>
      <input type=number name=days min=1 max=3650 value="{{ config.days }}" aria-label="Closed in last days">
      <span class=lbl>days</span>
    </span>
    <div class=preset-group role=group aria-label="Closed date presets">
      {% for d in [30, 60, 90, 180, 365] %}<a class=preset href="/closed?days={{d}}&amp;missing_tags={{ 1 if config.missing_tags else 0 }}&amp;photo_video_only={{ 1 if config.photo_video_only else 0 }}&amp;review_view={{ config.review_view }}" {% if config.days == d %}aria-current=page{% endif %}>{{d}}d</a>{% endfor %}
    </div>
  </div>
  <div class="panel-region region-groups">
    <fieldset class="filter-group scope-group">
      <legend class=group-lbl>Review Scope</legend>
      <div class=field><input type=hidden name=photo_video_only value=0><label for=closed-filter-photo-video><input type=checkbox id=closed-filter-photo-video name=photo_video_only value=1 {{ 'checked' if config.photo_video_only }}> Photo/Video Review Scope: ON</label></div>
      <p class=field-hint>Show only closed tickets whose subject indicates photos or video. Local display filter only — does not change Freshdesk or review state.</p>
    </fieldset>
    <fieldset class=filter-group>
      <legend class=group-lbl>Tags</legend>
      <div class=field><label for=closed-missing><input type=hidden name=missing_tags value=0><input type=checkbox id=closed-missing name=missing_tags value=1 {{ 'checked' if config.missing_tags }}> Missing Tags Only</label></div>
      <p class=field-hint>Show only closed tickets that have no tags.</p>
    </fieldset>
    <fieldset class=filter-group>
      <legend class=group-lbl>Review view</legend>
      <div class=field><label for=closed-review-view><select id=closed-review-view name=review_view>
        {% for v in ['active','completed','all'] %}<option value="{{ v }}" {{ 'selected' if config.review_view == v }}>{% if v == 'active' %}Active{% elif v == 'completed' %}Completed{% else %}All{% endif %}</option>{% endfor %}
      </select></label></div>
      <p class=field-hint>Active: Unreviewed · Opened / In Review · Needs Follow-Up · Completed: Resolved · N-A · No Action</p>
    </fieldset>
  </div>
  <div class="panel-region region-actions">
    <div class=action-buttons>
      <button type=submit class=apply>Apply Filters</button>
      <a class=reset href="/closed?days=60&amp;missing_tags=1&amp;photo_video_only=1&amp;review_view=active" role=button aria-label="Reset to defaults">Reset to Defaults</a>
    </div>
  </div>
</form>
<p class=review-note>Local review result only — does not change Freshdesk.</p>
{% if error %}<div class="banner err" role=alert>{{ error }}</div>{% elif result %}
<p class="filter-summary" role=status>{{ view_count }} of {{ result.unique_ticket_count }} unique closed tickets in {{ config.review_view }} view · {% if result.missing_tags_only %}Missing Tags Only{% else %}All tag states{% endif %} · {{ result.date_range[0] }} to {{ result.date_range[1] }} · {{ result.windows_planned|length }} date windows · {{ result.pages_requested|length }} pages · {% if result.complete %}<span class=complete>Complete</span>{% else %}<span class=incomplete>Results incomplete</span>{% endif %}</p>
{% for text in result.warnings %}<div class=incomplete role=status>{{ text }}</div>{% endfor %}{% for text in result.errors %}<div class=incomplete role=alert>{{ text }}</div>{% endfor %}
{% if closed_last_opened is not none %}{% if last_opened_rendered %}<p class=last-opened-bar><button type=button id=last-opened-jump aria-controls=closed-table>Jump to Last Opened</button></p>{% else %}<div class="banner" id=last-opened-hidden role=status>Last opened ticket is hidden by the current filters.</div>{% endif %}{% endif %}
{% if result.tickets %}<div class=tablewrap><table id=closed-table><caption class=visually-hidden>Closed ticket search results</caption><tr>
  <th scope=col>Ticket</th><th scope=col>Subject</th><th scope=col>Status</th>
  <th scope=col>Badges</th><th scope=col>Review</th><th scope=col>Closed</th>
  <th scope=col>Updated</th><th scope=col>Created</th><th scope=col>Tags</th>
</tr>{% for t in result.tickets %}<tr class="{{ t.row_class }}" data-ticket-id="{{ t.id }}"><td><a class="tid fd-link" href="{{ t.url }}" target=_blank rel="noopener noreferrer" data-ticket-id="{{ t.id }}" aria-label="Open ticket #{{ t.id }} in Freshdesk (new tab)">#{{ t.id }}</a></td><td><a class="sbj fd-link" href="{{ t.url }}" target=_blank rel="noopener noreferrer" data-ticket-id="{{ t.id }}" aria-label="Open subject of ticket #{{ t.id }} in Freshdesk (new tab)">{{ t.subject }}</a></td><td>{{ t.status_label }}</td><td><div class=badges>{% if t.last_opened %}<span class="badge b-last-opened">LAST OPENED</span>{% endif %}{% for kind, text, cls in t.badges %}<span class="badge {{ cls }}">{{ text }}</span>{% endfor %}</div></td><td><form class=rvform method=post action=/closed/api/review><input type=hidden name=csrf_token value="{{ csrf_token }}"><input type=hidden name=ticket_id value="{{ t.id }}"><input type=hidden name=days value="{{ config.days }}"><input type=hidden name=missing_tags value="{{ '1' if config.missing_tags else '0' }}"><input type=hidden name=photo_video_only value="{{ '1' if config.photo_video_only else '0' }}"><input type=hidden name=review_view value="{{ config.review_view }}"><select name=review_result aria-label="Review result for closed ticket {{ t.id }}" onchange="this.form.submit()">{% for s in review_states %}<option value="{{ s }}" {{ 'selected' if t.result == s }}>{{ s }}</option>{% endfor %}</select></form></td><td class=meta>{{ t.closed_display }}</td><td class=meta>{{ t.updated_display }}</td><td class=meta>{{ t.created_display }}</td><td>{% if t.tags %}{{ t.tags|join(', ') }}{% else %}<em style=color:#bbb>none</em>{% endif %}</td></tr>{% endfor %}</table></div>{% else %}<p class=closed-empty>No matching closed tickets were found in this view.</p>{% endif %}
<script>
var CSRF_TOKEN = {{ csrf_token_json | safe }};
var REVIEW_CLASS = {
  'Unreviewed': 'rv-unreviewed',
  'Opened / In Review': 'rv-opened',
  'Resolved': 'rv-resolved',
  'Not Applicable to Me': 'rv-na',
  'No Action Needed': 'rv-none',
  'Needs Follow-Up': 'rv-followup'
};
function badgeText(result) { return result === 'Opened / In Review' ? 'OPENED / IN REVIEW' : result; }
var toastEl = null;
function showError(msg) {
  if (!toastEl) {
    toastEl = document.createElement('div');
    toastEl.className = 'toast hidden';
    toastEl.setAttribute('role', 'status');
    document.body.appendChild(toastEl);
  }
  toastEl.textContent = msg;
  toastEl.classList.remove('hidden');
  clearTimeout(showError._t);
  showError._t = setTimeout(function () { toastEl.classList.add('hidden'); }, 6000);
}
// The Last Opened focus marker is derived server-side from the newest valid
// closed_last_opened_at and reported back by /closed/api/opened as
// last_opened_id. On a confirmed save we move the marker in the DOM without
// reloading (same mechanics as the /queue page; ids are per-page so both can
// coexist).
function moveLastOpened(newId) {
  if (typeof newId === 'undefined' || newId === null) { return; }
  ensureLastOpenedChrome();
  document.querySelectorAll('.b-last-opened').forEach(function (b) {
    b.parentNode.removeChild(b);
  });
  document.querySelectorAll('tr.rv-last-opened').forEach(function (r) {
    r.classList.remove('rv-last-opened');
  });
  var target = document.querySelector('tr[data-ticket-id="' + newId + '"]');
  var bar = document.querySelector('.last-opened-bar');
  var hidden = document.getElementById('last-opened-hidden');
  if (target) {
    target.classList.add('rv-last-opened');
    var badges = target.querySelector('.badges');
    if (badges) {
      var span = document.createElement('span');
      span.className = 'badge b-last-opened';
      span.textContent = 'LAST OPENED';
      badges.appendChild(span);
    }
    if (bar) { bar.style.display = ''; }
    if (hidden) { hidden.style.display = 'none'; }
  } else {
    if (bar) { bar.style.display = 'none'; }
    if (hidden) { hidden.style.display = ''; }
  }
}
function jumpToLastOpened() {
  var row = document.querySelector('tr.rv-last-opened');
  if (!row) { return; }
  row.scrollIntoView({behavior: 'smooth', block: 'center'});
  row.setAttribute('tabindex', '-1');
  row.focus({preventScroll: true});
  setTimeout(function () { row.removeAttribute('tabindex'); }, 1500);
}
function ensureLastOpenedChrome() {
  if (!document.querySelector('.last-opened-bar')) {
    var bar = document.createElement('p');
    bar.className = 'last-opened-bar';
    var btn = document.createElement('button');
    btn.type = 'button';
    btn.id = 'last-opened-jump';
    btn.setAttribute('aria-controls', 'closed-table');
    btn.textContent = 'Jump to Last Opened';
    bar.appendChild(btn);
    var anchor = document.querySelector('.tablewrap') || document.body;
    anchor.parentNode.insertBefore(bar, anchor);
  }
  if (!document.getElementById('last-opened-hidden')) {
    var div = document.createElement('div');
    div.className = 'banner';
    div.id = 'last-opened-hidden';
    div.setAttribute('role', 'status');
    div.textContent = 'Last opened ticket is hidden by the current filters.';
    var divAnchor = document.querySelector('.tablewrap') || document.body;
    divAnchor.parentNode.insertBefore(div, divAnchor);
  }
}
document.addEventListener('click', function (e) {
  if (e.target && e.target.id === 'last-opened-jump') { jumpToLastOpened(); }
});
// Ticket-number and subject links both carry data-ticket-id. Default
// navigation is never prevented: the Freshdesk tab opens synchronously and the
// local state request runs in the background.
document.querySelectorAll('a[data-ticket-id]').forEach(function (a) {
  a.addEventListener('click', function () {
    var tid = this.getAttribute('data-ticket-id');
    if (!tid) { return; }
    var row = this.closest('tr');
    fetch('/closed/api/opened', {
      method: 'POST',
      headers: {'Content-Type': 'application/json', 'X-CSRF-Token': CSRF_TOKEN},
      body: JSON.stringify({ticket_id: tid})
    }).then(function (r) { return r.json(); }).then(function (d) {
      if (d && d.ok && d.review_result) {
        var cls = REVIEW_CLASS[d.review_result] || 'rv-unreviewed';
        if (row) {
          var extra = [];
          row.classList.forEach(function (c) { if (c.indexOf('rv-') !== 0) { extra.push(c); } });
          row.className = extra.concat([cls]).join(' ');
        }
        var badge = row ? row.querySelector('.b-review') : null;
        if (badge) { badge.className = 'badge b-review ' + cls; badge.textContent = badgeText(d.review_result); }
        var sel = row ? row.querySelector('select[name=review_result]') : null;
        if (sel && sel.querySelector('option[value="' + d.review_result + '"]')) { sel.value = d.review_result; }
        moveLastOpened(d.last_opened_id);
      } else {
        showError('Could not save Opened / In Review state for #' + tid + ' (not saved).');
      }
    }).catch(function () {
      showError('Could not save Opened / In Review state for #' + tid + ' (not saved).');
    });
  });
});
// Filter controls: canonicalize on submit. The closed form carries days +
// missing_tags + review_view; unchecked checkboxes would be omitted by native
// GET submission, so rebuild one canonical query string from live state.
(function () {
  var form = document.querySelector('form.controls');
  if (!form) { return; }
  function normDays(raw) {
    var v = String(raw == null ? '' : raw).trim();
    if (!/^[0-9]+$/.test(v)) { return 60; }
    var n = parseInt(v, 10);
    return (n >= 1 && n <= 3650) ? n : 60;
  }
  function normView(v) { return (v === 'completed' || v === 'all') ? v : 'active'; }
  form.addEventListener('submit', function (e) {
    try {
      var params = {};
      var chk = form.querySelector('input[name=missing_tags]');
      params['missing_tags'] = chk && chk.checked ? '1' : '0';
      var daysEl = form.querySelector('input[name=days]');
      params['days'] = String(normDays(daysEl ? daysEl.value : null));
      var viewEl = form.querySelector('select[name=review_view]');
      params['review_view'] = normView(viewEl ? viewEl.value : 'active');
      var parts = [];
      ['missing_tags', 'days', 'review_view'].forEach(function (k) {
        parts.push(encodeURIComponent(k) + '=' + encodeURIComponent(params[k]));
      });
      e.preventDefault();
      window.location.href = '/closed?' + parts.join('&');
    } catch (err) {
      // On any unexpected error, fall back to native submission.
    }
  });
  function syncControlsFromURL() {
    var q = new URLSearchParams(window.location.search);
    var mt = q.get('missing_tags') !== '0';
    var mtEl = form.querySelector('input[name=missing_tags][value=1]');
    if (mtEl) { mtEl.checked = mt; }
    var d = q.get('days');
    var daysEl = form.querySelector('input[name=days]');
    if (daysEl) { daysEl.value = (/^[0-9]+$/.test(d || '') && parseInt(d, 10) >= 1) ? d : 60; }
    var v = q.get('review_view');
    var viewEl = form.querySelector('select[name=review_view]');
    if (viewEl) { viewEl.value = (v === 'completed' || v === 'all') ? v : 'active'; }
  }
  window.addEventListener('pageshow', function () {
    try { syncControlsFromURL(); } catch (err) {}
  });
})();
(function () {
  var start = document.getElementById('closed-refresh');
  if (!start) { return; }
  var cancel = document.getElementById('closed-cancel');
  var output = document.getElementById('closed-refresh-status');
  var terminalSeen = false;
  function encode(obj) { return Object.keys(obj).map(function(k){ return encodeURIComponent(k)+'='+encodeURIComponent(obj[k]); }).join('&'); }
  function render(s) {
    var p = s.progress || {};
    var running = s.state === 'running';
    start.disabled = running;
    cancel.hidden = !running;
    var msg = s.message || '';
    if (running) {
      msg += ' Page ' + (p.page || p.pages_completed || 0) + ' · ' + (p.rows_received || 0) + ' rows · ' + (p.unique_tickets || 0) + ' unique';
      if (p.rate_limit_remaining != null) { msg += ' · rate-limit remaining ' + p.rate_limit_remaining; }
      if (p.waiting_seconds) { msg += ' · waiting ' + p.waiting_seconds + ' seconds'; }
    }
    output.textContent = msg;
    if (running) { window.setTimeout(poll, 1000); }
    else if ((s.state === 'success' || s.state === 'succeeded') && !terminalSeen) { terminalSeen = true; window.location.reload(); }
  }
  function poll() { fetch('/closed/api/refresh/status').then(function(r){return r.json();}).then(render).catch(function(){ output.textContent='Unable to read refresh status.'; }); }
  start.addEventListener('click', function () {
    terminalSeen = false;
    var days = document.querySelector('form.controls input[name=days]');
    fetch('/closed/api/refresh', {method:'POST', headers:{'Content-Type':'application/x-www-form-urlencoded'}, body:encode({csrf_token:CSRF_TOKEN, days:days ? days.value : '60'})})
      .then(function(r){return r.json();}).then(render);
  });
  cancel.addEventListener('click', function () {
    fetch('/closed/api/refresh/cancel', {method:'POST', headers:{'Content-Type':'application/x-www-form-urlencoded'}, body:encode({csrf_token:CSRF_TOKEN})})
      .then(function(r){return r.json();}).then(function(s){output.textContent=s.message || ''; poll();});
  });

  // No page-load poll: opening or reloading /closed performs zero requests.
  // Status polling starts only from the explicit Refresh/Cancel actions above.
})();
</script>
{% endif %}</body></html>
"""


def _closed_render(**kwargs):
    """Render CLOSED_HTML with the shared context merged in."""
    ctx = dict(kwargs)
    ctx.setdefault("offline", True)
    ctx.setdefault("shared_css", _SHARED_CSS)
    ctx.setdefault("nav", _nav_html("closed"))
    ctx.setdefault("status_label", status_label)
    ctx.setdefault("csrf_token", get_csrf_token())
    ctx.setdefault("csrf_token_json", json.dumps(ctx["csrf_token"]))
    ctx.setdefault("review_states", REVIEW_STATES)
    ctx.setdefault("flash", None)
    ctx.setdefault("view_count", 0)
    ctx.setdefault("closed_last_opened", None)
    ctx.setdefault("last_opened_rendered", False)
    ctx.setdefault("live_cache", closed_live.load_cache() if not ctx["offline"] else None)
    return render_template_string(CLOSED_HTML, **ctx)


def _queue_render(**kwargs):
    """Render QUEUE_HTML with the shared context merged in."""
    ctx = dict(kwargs)
    cfg = ctx.get("config") or dict(DEFAULT_FILTERS)
    cfg.setdefault("workflow_tab", "main")
    ctx.setdefault("config", cfg)
    ctx.setdefault("shared_css", _SHARED_CSS)
    ctx.setdefault("nav", _nav_html("queue"))
    ctx.setdefault("csrf_token", get_csrf_token())
    ctx.setdefault("flash", None)
    token = ctx["csrf_token"]
    ctx.setdefault("csrf_token_json", json.dumps(token))
    ctx.setdefault("review_states", REVIEW_STATES)
    ctx.setdefault("days_min", DAYS_MIN)
    ctx.setdefault("days_max", DAYS_MAX)
    ctx.setdefault("preset_urls", {d: filter_query_string(dict(cfg, days=d)) for d in (7, 14, 30, 60, 90)})
    ctx.setdefault("active_summary", filter_summary_text(cfg))
    ctx.setdefault("filter_query_string", filter_query_string)
    ctx.setdefault("workflow_counts", {tab: 0 for tab in WORKFLOW_TABS})
    ctx.setdefault("workflow_labels", WORKFLOW_LABELS)
    ctx.setdefault("queue_scope_counts", {"main": 0, "triage": 0})
    ctx.setdefault("live_cache_missing", False)
    ctx.setdefault("last_refresh_display", "Never")
    ctx.setdefault("cache_coverage_display", "Unknown")
    ctx.setdefault("cached_ticket_count", 0)
    ctx.setdefault("last_refresh_mode_display", "No baseline")
    ctx.setdefault("auto_refresh", auto_refresh_status())
    return render_template_string(QUEUE_HTML, **ctx)


if __name__ == "__main__":
    initialize_live_auto_refresh()
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "5050"))
    app.run(host=resolve_bind_host(host), port=port, debug=False)
